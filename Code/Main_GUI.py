import csv
import os
import tkinter
from tkinter import *
import re
from sklearn.preprocessing import StandardScaler
import numpy as np
from scipy import stats
from tkinter.ttk import Combobox
from tkinter import Tk, filedialog
import pandas as pd
from sklearn import svm, datasets
from sklearn.metrics import roc_curve, auc
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import label_binarize
from sklearn.multiclass import OneVsRestClassifier
from scipy import interp
from tkinter import messagebox
import random
import openpyxl
from matplotlib import pyplot as plt
from openpyxl.chart import ScatterChart, Reference, BarChart3D
from prettytable import PrettyTable
from keras.models import Sequential
from keras.layers import Dense, Embedding
from keras.layers import LSTM
from keras.callbacks import EarlyStopping
import math
import time
import numpy as np
import torch
from torch import nn
from torch.nn.parameter import Parameter
from torch.nn.init import normal_
import torch.nn.functional as F
from imblearn.over_sampling import ADASYN
import pandas as pd


from IDS import config as cfg
from IDS.Clustering.ExistingFCM import ExistingFCM
from IDS.Clustering.ProposedHeWaFCM import ProposedHeWaFCM
from IDS.Clustering.ExistingBirch import ExistingBirch
from IDS.Clustering.ExistingKmedoid import ExistingKMedoid
from IDS.Clustering.ExistingKMeans import ExistingKMeans
from IDS.Classification.Existing_DNN import Existing_DNN
from IDS.Classification.Existing_LSTM import Existing_LSTM
from IDS.Classification.Existing_RNN import Existing_RNN
from IDS.Classification.Existing_DBN import Existing_DBN
from IDS.Classification.Proposed_FedTL_SRSKLSTM import Proposed_FedTL_SRSKLSTM


class Main_GUI:
    boolfile = False

    bool_ddr = False
    bool_mvi = False
    bool_normalization = False
    bool_numeralization = False
    bool_feature_extraction = False
    bool_feature_selection = False
    bool_dataset_splitting = False
    bool_training =False
    bool_testing = False
    bool_sensor_data_testing = False
    bool_map_reduce=False
    bool_ts_feature_extraction=False
    bool_feature_similarity=False
    bool_sink_node=False

    bool_tr_file_read = False

    bool_initalize_nodes = False
    bool_ch_selection = False
    bool_clustering = False

    tdata = []
    trdata = []
    tsdata = []
    trddrdata = []
    trmvidata = []
    tsmvidata = []
    trstrval = []
    trnumdata = []
    tsnumdata = []
    trnomdata = []
    tsnomdata = []
    trattributes = []
    tsattributes = []
    tscls=[]
    sim_fe = []
    du_tscls=[]
    baldata=[]

    sensingdata = []

    position = []
    energy = []
    cfg.node_name = []
    trainingsize = 80
    testingsize = 20
    cfg.iteration = 10

    ipdata = []
    iptrfeatures=[]

    iptrdata = []
    iptsdata = []

    iptrcls = []
    iptscls = []
    res = []

    def __init__(self, root):
        self.file_path = StringVar()
        self.noofnodes = StringVar()
        self.menu = StringVar()

        self.LARGE_FONT = ("Algerian", 16)
        self.text_font = ("Constantia", 15)
        self.text_font1 = ("Constantia", 10)

        self.frame_font = ("", 9)
        self.frame_process_res_font = ("", 12)
        self.root = root
        self.node_value = StringVar()

        label_heading = tkinter.Label(root, text="SIGNIFICANT FEATURE SIMILARITY EVALUATION BASED IDS USING JSKWT AND FedTL-SRSKLSTM", fg="deep pink", bg="wheat", font=self.LARGE_FONT)
        label_heading.place(x=100, y=0)

        self.label_testing = LabelFrame(root, text="Testing", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_testing.place(x=10, y=165, width=170, height=500)
        self.label_node_init = LabelFrame(root, text="Network Initialization", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_node_init.place(x=20, y=185, width=150, height=85)
        self.btn_initalize_nodes = Button(root, text="Node initialization", bg="deep sky blue", fg="#fff",font=self.text_font1, width=5, command=self.initalize_nodes)
        self.btn_initalize_nodes.place( x = 35, y = 240, width = 120)

        self.cmb_nodes = Combobox(root, width=20, height=30, textvariable=self.noofnodes)
        self.cmb_nodes["values"] = ("50", "100", "150", "200", "250")
        self.cmb_nodes.set("Select no. of Nodes")
        self.cmb_nodes.configure(state='readonly')
        self.cmb_nodes.grid(column=1, row=5)
        self.cmb_nodes.place(x=35, y=208,width=120,height = 25)
        #
        self.label_clustering_forms = LabelFrame(root, text="Clustering Forms", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_clustering_forms.place(x=20, y=270, width=150, height=380)
        self.label_ch_selection = LabelFrame(root, text="CH Selection", bg="wheat", fg="mediumblue",font=self.frame_font)
        self.label_ch_selection.place(x=35, y=290, width=100, height=55)
        self.btn_ch_selection = Button(root, text="Proceed", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command=self.ch_selection)
        self.btn_ch_selection.place(x=45, y=310, width=70)
        self.label_clustering = LabelFrame(root, text="Clustering", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_clustering.place(x=35, y=350, width=100, height=55)
        self.btn_clustering = Button(root, text="Proceed", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.clustering)
        self.btn_clustering.place(x=45, y=370, width=70)
        self.label_nw_traffic_analy = LabelFrame(root, text="N/W Traffic Analyzer", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_nw_traffic_analy.place(x=35, y=410, width=120, height=55)
        self.btn_map_reduce= Button(root, text="Map Reduce", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.map_reduce)
        self.btn_map_reduce.place(x=45, y=430, width=80)
        self.label_feature_ext = LabelFrame(root, text="Feature Extraction", bg="wheat", fg="mediumblue",font=self.frame_font)
        self.label_feature_ext.place(x=35, y=470, width=120, height=55)
        self.btn_ts_feature_extraction = Button(root, text="Proceed", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command=self.ts_feature_extraction)
        self.btn_ts_feature_extraction.place(x=45, y=490, width=70)
        self.label_feature_similarity = LabelFrame(root, text="Similarity Estimation", bg="wheat", fg="mediumblue",font=self.frame_font)
        self.label_feature_similarity.place(x=35, y=530, width=125, height=55)
        self.btn_feature_similarity = Button(root, text="Proceed", bg="deep sky blue", fg="#fff", font=self.text_font1,width=5, command=self.feature_similarity)
        self.btn_feature_similarity.place(x=45, y=550, width=70)
        #
        self.btn_sink_node = Button(root, text="Model\nUpdated", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command=self.sink_node)
        self.btn_sink_node.place(x=25, y=590, width=60, height=50)

        self.btn_classifier = Button(root, text="Classifier", bg="deep sky blue", fg="#fff", font=self.text_font1,width=5, command=self.ts_classifier)
        self.btn_classifier.place(x=105, y=590, width=60, height=50)
        ##########Training#############

        self.label_training = LabelFrame(root, text="Training", bg="wheat", fg="mediumblue",
                                                   font=self.frame_font)
        self.label_training.place(x=10, y=35, width=830, height=120)
        self.label_crop_field_dataset = LabelFrame(root, text="Training Dataset", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_crop_field_dataset.place(x=20, y=55, width=165, height=90)
        self.selected_dataset = StringVar()
        dname = ['NSL-KDD', 'KDD CUP-99']
        self.dataset_cb = Combobox(root, values=dname,state="readonly", textvariable=self.selected_dataset)
        self.dataset_cb.place(x=30, y=80, width=150, height=30)


        self.btn_tr_dataset = Button(root, text="Read Data", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command = self.tr_dataset_file_browse)
        self.btn_tr_dataset.place(x=35, y=115,width=120)

        self.label_preprocessing = LabelFrame(root, text="Preprocessing", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_preprocessing.place(x=190, y=55, width=210, height=90)
        self.btn_DDR = Button(root, text="De Duplication", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.DDR)
        self.btn_DDR.place(x=245, y=75,width=100)
        self.btn_Numeralization = Button(root, text="Numeralization", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.Numeralization)
        self.btn_Numeralization.place(x=195, y=105,width=90)
        self.btn_Normalization = Button(root, text="Normalization", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.Normalization)
        self.btn_Normalization.place(x=305, y=105,width=90)

        self.btn_data_balancing = Button(root, text="Data Balancing", bg="deep sky blue", fg="#fff", font=self.text_font1,width=5, command=self.data_balancing)
        self.btn_data_balancing.place(x=415, y=70, width=105)

        # self.label_feature_extraction = LabelFrame(root, text="Feature Extraction", bg="wheat", fg="mediumblue",font=self.frame_font)
        # self.label_feature_extraction.place(x=405, y=55, width=120, height=55)
        self.btn_feature_extraction = Button(root, text="Feature Extraction", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.feature_extraction)
        self.btn_feature_extraction.place(x=415, y=105, width=105)

        self.label_feature_evaluation = LabelFrame(root, text="Feature Evaluation", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_feature_evaluation.place(x=535, y=55, width=120, height=55)
        self.btn_feature_evaluation = Button(root, text="Proceed", bg="deep sky blue", fg="#fff", font=self.text_font1,width=5, command=self.feature_evaluation)
        self.btn_feature_evaluation.place(x=545, y=75, width=100)

        self.label_dimensionality_reduction = LabelFrame(root, text="Dimensionality Reduction", bg="wheat", fg="mediumblue",font=self.frame_font)
        self.label_dimensionality_reduction.place(x=665, y=55, width=160, height=55)
        self.btn_dimensionality_reduction = Button(root, text="Proceed", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command=self.dimensionality_reduction)
        self.btn_dimensionality_reduction.place(x=695, y=75, width=100)
        #
        self.label_dataset_splitting = LabelFrame(root, text="Classification", bg="wheat", fg="mediumblue",font=self.frame_font)
        self.label_dataset_splitting.place(x=855, y=35, width=150, height=120)
        self.btn_dataset_splitting = Button(root, text="Dataset Spliting", bg="deep sky blue", fg="#fff", font=self.text_font1,width=5, command=self.dataset_splitting)
        self.btn_dataset_splitting.place(x=875, y=55, width=100)
        self.btn_training = Button(root, text="Training", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command=self.training)
        self.btn_training.place(x=875, y=88, width=100)
        self.btn_testing = Button(root, text="Testing", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.testing)
        self.btn_testing.place(x=875, y=120, width=100)


        self.label_result_graphs = LabelFrame(root, text="Result & Graphs", bg="wheat", fg="mediumblue", font=self.frame_font)
        self.label_result_graphs.place(x=1015, y=80, width=105, height=55)
        self.btn_result_graphs = Button(root, text="Proceed", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5,command=self.result_graphs)
        self.btn_result_graphs.place(x=1025, y=100, width=80)

        self.btn_clear = Button(root, text="Clear", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command=self.clear)
        self.btn_clear.place(x=900, y=630, width=90,height=30)

        self.btn_exit = Button(root, text="Exit", bg="deep sky blue", fg="#fff", font=self.text_font1, width=5, command=self.exit)
        self.btn_exit.place(x=1000, y=630, width=90,height=30)

        # Horizontal (x) Scroll bar
        self.xscrollbar = Scrollbar(root, orient=HORIZONTAL)
        self.xscrollbar.pack(side=BOTTOM, fill=X)
        # Vertical (y) Scroll Bar
        self.yscrollbar = Scrollbar(root)
        self.yscrollbar.pack(side=RIGHT, fill=Y)

        self.label_output_frame = LabelFrame(root, text="Process Window", bg="wheat", fg="#0000FF", font=self.frame_process_res_font)
        self.label_output_frame.place(x=200, y=165, width=450, height=460)
        # Text Widget
        self.data_textarea_process = Text(root, wrap=WORD, xscrollcommand=self.xscrollbar.set, yscrollcommand=self.yscrollbar.set)
        self.data_textarea_process.pack()
        # Configure the scrollbars
        self.xscrollbar.config(command=self.data_textarea_process.xview)
        self.yscrollbar.config(command=self.data_textarea_process.yview)
        self.data_textarea_process.place(x=210, y=185, width=430, height=430)
        self.data_textarea_process.configure(state="disabled")

        self.label_output_frame = LabelFrame(root, text="Result Window", bg="wheat", fg="#0000FF", font=self.frame_process_res_font)
        self.label_output_frame.place(x=670, y=165, width=450, height=460)
        # Text Widget
        self.data_textarea_result = Text(root, wrap=WORD, xscrollcommand=self.xscrollbar.set, yscrollcommand=self.yscrollbar.set)
        self.data_textarea_result.pack()
        # Configure the scrollbars
        self.xscrollbar.config(command=self.data_textarea_result.xview)
        self.yscrollbar.config(command=self.data_textarea_result.yview)
        self.data_textarea_result.place(x=680, y=185, width=430, height=430)
        self.data_textarea_result.configure(state="disabled")

    def tr_dataset_file_browse(self):
        strval = self.dataset_cb.get()
        cfg.data_name = self.dataset_cb.get()
        if len(strval) > 0:
            self.bool_tr_file_read = True
            if self.selected_dataset.get() == "NSL-KDD":
                print("\nSelected Dataset Name : " + str(self.selected_dataset.get()))

                self.data_textarea_process.insert(INSERT, "\nSelected Dataset Name : " + str(self.selected_dataset.get()))

                with open("..\\Dataset\\KDDTrain+.txt") as f:
                    lines = f.readlines()
                    cnt = 0
                    temp=[]
                    for x in range(len(lines)):
                        strval = str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip()
                        lpos = strval.rindex(",")
                        self.ipdata.append(strval[0:lpos])

                        # if cnt >= 50000:
                        #     break
            else:
                print("\nSelected Dataset Name : " + str(self.selected_dataset.get()))
                self.data_textarea_process.insert(INSERT, "\nSelected Dataset Name : " + str(self.selected_dataset.get()))

                with open("..\\Dataset\\kddcup.data.corrected") as f:
                    lines = f.readlines()
                    cnt = 0
                    for x in range(len(lines)):
                        # cnt += 1
                        self.ipdata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())

                        # if cnt >= 200000:
                        #     break

            for x in range(len(self.ipdata)):
                print(self.ipdata[x])

            print("\n Total No. of Data: "+str(len(self.ipdata)))
            print("\nDataset was selected successfully...")
            messagebox.showinfo("Information Message", "Dataset was selected successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nDataset was selected successfully...")

            self.data_textarea_process.configure(state='disabled')
            self.data_textarea_result.configure(state='disabled')
            self.btn_tr_dataset.configure(state="disabled")
        else:
            messagebox.showinfo("Information Message", "Please Select the Dataset...")

    def DDR(self):
        if self.bool_tr_file_read:
            self.bool_ddr = True
            self.data_textarea_process.configure(state="normal")

            print("\nPre-processing")
            print("================")
            print(" De Duplication")
            print("---------------------")

            self.data_textarea_process.insert(INSERT, "\n\nPre-processing")
            self.data_textarea_process.insert(INSERT, "\n================")
            self.data_textarea_process.insert(INSERT, "\nDe Duplication")
            self.data_textarea_process.insert(INSERT, "\n---------------------")

            print("Total No. of Training Data: " + str(len(self.ipdata)))
            self.data_textarea_process.insert(INSERT, "\nTotal No. of Training Data: " + str(len(self.ipdata)))
            temp = (list(set(self.ipdata)))
            temp1=[]
            for x in range(len(temp)):
                cls_data = temp[x].strip(',').split(',')[-1]

                sen_node = temp[x].strip(',').split(',')[:-1]
                self.du_tscls.append(cls_data)
                self.trddrdata.append(sen_node)

            # for x in range(len(self.trddrdata)):
            #     print(self.trddrdata[x])

            print("Total no. of Training Data after repeated data removal : " + str(len(self.trddrdata)))
            self.data_textarea_process.insert(INSERT,"\nTotal no. of Training Data after repeated data removal : " + str(len(self.trddrdata)))



            messagebox.showinfo("Info Message", "Training De Duplication was done successfully...")
            print("\nTraining De Duplication was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nTraining De Duplication was done successfully...")

            self.btn_DDR.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please read the Dataset First ...")

    def chStringAvailability(self, arrval):
        strval = []
        for x in range(len(arrval)):
            for y in range(len(arrval[x])):
                try:
                    if float(arrval[x][y]):
                        pass
                except:
                    strval.append(arrval[x][y])
                    pass

        return strval

    def findNumeralization(self, sval, arrval):
        numval = []
        for x in range(len(arrval)):
            temp = []
            for y in range(len(arrval[x])):
                if sval.__contains__(arrval[x][y]):
                    temp.append(float(sval.index(arrval[x][y])))
                else:
                    temp.append(float(arrval[x][y]))
            numval.append(temp)
        return numval

    def Numeralization(self):    
            self.bool_numeralization=True
            self.data_textarea_process.configure(state="normal")

            print("\nNumeralization")
            print("----------------")
            self.data_textarea_process.insert(INSERT, "\n\nNumeralization")
            self.data_textarea_process.insert(INSERT, "\n----------------")
            if not os.path.exists("..\\Result\\Training\\"):
                os.makedirs("..\\Result\\Training\\")
            if self.selected_dataset.get() == "NSL-KDD":

                    unrepeat_data = []
                    count = 0
                    for x in self.trddrdata:
                        for sdata in x:
                            if type(sdata) is str:
                                if sdata not in unrepeat_data:
                                    unrepeat_data.append(sdata)

                    count1 = 0

                    for x in self.trddrdata:
                        smlst = []
                        for sdata in x:
                            if sdata in unrepeat_data:
                                smlst.append(unrepeat_data.index(sdata))
                            else:
                                smlst.append(sdata)
                        self.trnumdata.append(smlst)
                        count1 += 1

                    df = pd.DataFrame(self.trnumdata)
                    # print(df)
                    df.to_csv('..\\Result\\Training\\Numeralization_NSL_KDD.csv', index=False)
            else:
                    unrepeat_data = []
                    count = 0
                    for x in self.trddrdata:
                        for sdata in x:
                            if type(sdata) is str:
                                if sdata not in unrepeat_data:
                                    unrepeat_data.append(sdata)

                    count1 = 0

                    for x in self.trddrdata:
                        smlst = []
                        for sdata in x:
                            if sdata in unrepeat_data:
                                smlst.append(unrepeat_data.index(sdata))
                            else:
                                smlst.append(sdata)
                        self.trnumdata.append(smlst)
                        count1 += 1

                    df = pd.DataFrame(self.trnumdata)
                    # print(df)
                    df.to_csv('..\\Result\\Training\\Numeralization_KDD_CUP_99.csv', index=False)
           
            for x in range(len(self.trnumdata)):
                print(self.trnumdata[x])

            messagebox.showinfo("Info Message", "Numeralization was done successfully...")
            print("\nNumeralization was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nNumeralization was done successfully...")

            self.btn_Numeralization.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
        
    def findNormalization(self, arrval):
        # Linear Scaling
        # x ′ = (x − x m i n) / (x m a x − x m i n)

        nomval = []
        for x in range(len(arrval)):
            temp = []
            vmin = float(min(arrval[x]))
            vmax = float(max(arrval[x]))
            for y in range(len(arrval[x])):
                vv = float(arrval[x][y])
                val = ((vv - vmin) / (vmax - vmin))
                temp.append(val)
            nomval.append(temp)
        return nomval

    def Normalization(self):
        if self.bool_numeralization:
            self.bool_normalization=True
            self.data_textarea_process.configure(state="normal")

            print("\nNormalization")
            print("---------------")
            self.data_textarea_process.insert(INSERT, "\n\nNormalization")
            self.data_textarea_process.insert(INSERT, "\n---------------")
            if not os.path.exists("..\\Result\\Training\\"):
                os.makedirs("..\\Result\\Training\\")
            if self.selected_dataset.get() == "NSL-KDD":

                    self.trnomdata = self.findNormalization(self.trnumdata)
                    df = pd.DataFrame(self.trnomdata)

                    df2 = df.assign(cls=self.du_tscls)
                    print(df2)
                    df2.to_csv('..//Result//Training//Normalization_NSL_KDD.csv', index=False)


            else:
                    self.trnomdata = self.findNormalization(self.trnumdata)
                    df = pd.DataFrame(self.trnomdata)

                    df2 = df.assign(cls=self.du_tscls)
                    print(df2)
                    df2.to_csv('..//Result//Training//Normalization_KDD_CUP_99.csv', index=False)

            for x in range(len(self.trnomdata) - 1):
                print(self.trnomdata[x])


            messagebox.showinfo("Info Message", "Normalization was done successfully...")
            print("\nNormalization was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nNormalization was done successfully...")

            self.btn_Normalization.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please done the Numeralization data First ...")

    def data_balancing(self):
        # if self.bool_normalization:

            print("\nData Balancing")
            print("---------------")
            self.data_textarea_process.insert(INSERT, "\n\nData Balancing")
            self.data_textarea_process.insert(INSERT, "\n---------------")
            if self.selected_dataset.get() == "NSL-KDD":
                if os.path.exists("..//Result//Training//Balanced_data_NSL_KDD.csv"):
                    ipdata = []

                    with open('../Result/Training/Normalization_NSL_KDD.csv') as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):  # cnt += 1

                            ipdata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(ipdata)):
                        sen_node = ipdata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))

                    test_list = list(set(temp))



                    with open("..//Result//Training//Balanced_data_NSL_KDD.csv") as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):
                            self.baldata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(self.baldata)):
                        sen_node = self.baldata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))
                    bal_list = list(set(temp))

                    print("\nNumber of data before balancing:" + str(len(ipdata)))
                    print("Number of data after balancing:" + str(len(self.baldata)))

                else:
                    ipdata = []
                    with open("..\\Result\\Training\\Normalization_NSL_KDD.csv") as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):
                            ipdata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(ipdata)):
                        sen_node = ipdata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    # print(temp)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))

                    test_list = list(set(temp))
                    gr_data = []
                    lo_data = []

                    for x in range(len(test_list)):
                        if temp.count(test_list[x]) > 1500:
                            gr_data.append(test_list[x])
                        elif temp.count(test_list[x]) < 1500:
                            lo_data.append(test_list[x])
                    df = pd.read_csv("..\\Result\\Training\\Normalization_NSL_KDD.csv", index_col=0)
                    df1 = df[df["cls"].str.contains('|'.join(lo_data)) == False]
                    df = df[df["cls"].str.contains('|'.join(gr_data)) == False]
                    classes = df.values[:, -1]
                    data = df.iloc[:, 0:-1]
                    adasyn = ADASYN(sampling_strategy='not majority', random_state=8, n_neighbors=1)
                    #
                    balanced_data, new_classes = adasyn.fit_resample(data, classes)
                    new_cls = pd.DataFrame(new_classes, columns=["cls"])
                    result = pd.concat([balanced_data, new_cls], axis=1, join='inner')
                    frames = [df1, result]
                    result1 = pd.concat(frames)
                    result1.to_csv('..//Result//Training//Balanced_data_NSL_KDD.csv', index=False)
                    # baldata = []
                    with open("..//Result//Training//Balanced_data_NSL_KDD.csv") as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):
                            self.baldata.append(
                                str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(self.baldata)):
                        sen_node = self.baldata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))
                    bal_list = list(set(temp))

                    print("\nNumber of data before balancing:" + str(len(ipdata)))
                    print("Number of data after balancing:" + str(len(self.baldata)))
            else:
                if os.path.exists("..//Result//Training//Balanced_KDD_CUP_99.csv"):
                    ipdata = []

                    with open('..//Result//Training//Normalization_KDD_CUP_99.csv') as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):  # cnt += 1

                            ipdata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(ipdata)):
                        sen_node = ipdata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))

                    test_list = list(set(temp))


                    baldata = []
                    with open("..//Result//Training//Balanced_KDD_CUP_99.csv") as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):
                            baldata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(baldata)):
                        sen_node = baldata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))
                    bal_list = list(set(temp))

                    print("\nNumber of data before balancing:" + str(len(ipdata)))
                    print("Number of data after balancing:" + str(len(baldata)))

                else:
                    ipdata = []
                    with open("..\\Result\\Training\\Normalization_KDD_CUP_99.csv") as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):
                            ipdata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(ipdata)):
                        sen_node = ipdata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    # print(temp)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))

                    test_list = list(set(temp))
                    gr_data = []
                    lo_data = []

                    for x in range(len(test_list)):
                        if temp.count(test_list[x]) > 3500:
                            gr_data.append(test_list[x])
                        elif temp.count(test_list[x]) < 3500:
                            lo_data.append(test_list[x])
                    df = pd.read_csv("..\\Result\\Training\\Normalization_KDD_CUP_99.csv", index_col=0)
                    df1 = df[df["cls"].str.contains('|'.join(lo_data)) == False]
                    df = df[df["cls"].str.contains('|'.join(gr_data)) == False]
                    classes = df.values[:, -1]
                    data = df.iloc[:, 0:-1]
                    adasyn = ADASYN(sampling_strategy='not majority', random_state=8, n_neighbors=1)
                    #
                    balanced_data, new_classes = adasyn.fit_resample(data, classes)
                    new_cls = pd.DataFrame(new_classes, columns=["cls"])
                    result = pd.concat([balanced_data, new_cls], axis=1, join='inner')
                    frames = [df1, result]
                    result1 = pd.concat(frames)

                    result1.to_csv('..//Result//Training//Balanced_KDD_CUP_99.csv', index=False)
                    # baldata = []
                    with open("..//Result//Training//Balanced_KDD_CUP_99.csv") as f:
                        lines = f.readlines()
                        cnt = 0
                        for x in range(len(lines)):
                            self.baldata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    temp = []
                    for x in range(len(self.baldata)):
                        sen_node = self.baldata[x].strip(',').split(',')[-1]
                        temp.append(sen_node)
                    remove = "cls"
                    if remove in temp:
                        temp.pop(temp.index(remove))
                    bal_list = list(set(temp))

                    print("Number of data before balancing:" + str(len(ipdata)))
                    print("Number of data after balancing:" + str(len(self.baldata)))

            messagebox.showinfo("Info Message", "Data Balancing was done successfully...")
            print("\nData Balancing was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nData Balancing was done successfully...")
            self.btn_data_balancing.configure(state="disable")


        # else:
        #     messagebox.showinfo("Info Message", "Please perform Normalization first...")



    def feature_extraction(self):
        if self.bool_numeralization:
            self.bool_feature_extraction = True
            self.data_textarea_process.configure(state="normal")
            print("\nFeature Extraction")
            print("======================")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Extraction")
            self.data_textarea_process.insert(INSERT, "\n======================")
            self.data_textarea_result.insert(INSERT, "\n\nFeature Extraction")
            self.data_textarea_result.insert(INSERT, "\n======================")
            if self.selected_dataset.get() == "NSL-KDD":
                with open(str("..\\Dataset\\Attribute_NSL_KDD.txt")) as f:
                    lines = f.readlines()
                    for sub in lines:
                        cfg.trattributes.append(re.sub('\n', '', sub))
                        len(cfg.trattributes)
                    for x in range(len(cfg.trattributes) - cfg.tr_data):
                        cfg.val_FE.append(cfg.trattributes[x])
            else:
                with open(str("..\\Dataset\\Attribute_KDD CUP-99.txt")) as f:
                    lines = f.readlines()
                    for sub in lines:
                        cfg.trattributes.append(re.sub('\n', '', sub))
                        len(cfg.trattributes)
                    for x in range(len(cfg.trattributes) - cfg.tr_data):
                        cfg.val_FE.append(cfg.trattributes[x])


            print("Total no. of Attributes : " + str(len(cfg.trattributes)))
            self.data_textarea_result.insert(INSERT, "\nTotal no. of Attributes : " + str(len(cfg.trattributes)))

            print("\nFeatures are...")
            print("---------------------")
            print(cfg.trattributes)

            self.data_textarea_result.insert(INSERT, "\n\nFeatures are...")
            self.data_textarea_result.insert(INSERT, "\n---------------------")
            self.data_textarea_result.insert(INSERT, "\n" + str(cfg.trattributes))

            messagebox.showinfo("Info Message", "Feature Extraction was done successfully...")
            print("\nFeature Extraction was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Extraction was done successfully...")

            self.btn_feature_extraction.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please done the Normalization data First ...")

    def feature_evaluation(self):
        if self.bool_feature_extraction:
            self.bool_feature_selection = True
            self.data_textarea_process.configure(state="normal")
            self.data_textarea_result.configure(state="normal")
            print("\nFeature Evaluation")
            print("======================")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Evaluation")
            self.data_textarea_process.insert(INSERT, "\n===================")

            if not os.path.exists("../Result\\"):
                os.mkdir("../Result\\")

            arr_t = np.array(self.trnomdata).T

            val_FE=[]
            for x in range(len(arr_t)-1):
                temp = []
                result = stats.kruskal(arr_t[x], arr_t[2])
                if result.pvalue>=0.05:
                    temp.append(result.pvalue)
                    temp.append(cfg.trattributes[x])
                    val_FE.append(temp)
            with open('..\\Result\\Significant Feature.txt', 'w') as f:
                for x in range(len(cfg.val_FE)):
                    f.write(cfg.val_FE[x]+ "\n")
            with open('..\\Result\\validate Feature.txt', 'w') as f:
                for x in range(len(cfg.val_FE)):
                    f.write(cfg.val_FE[x]+ "\n")

            print("\n No. of Evaluate Features: "+str(len(cfg.val_FE)))
            print("\n Evaluate Features: " + str(cfg.val_FE))
            messagebox.showinfo("Info Message", "Feature Evaluation was done successfully...")
            print("\nFeature Evaluation was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Evaluation was done successfully...")

            self.btn_feature_evaluation.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please done the Feature extraction First ...")


    def dimensionality_reduction(self):
        # if self.bool_feature_selection:
            self.data_textarea_process.configure(state="normal")
            self.data_textarea_result.configure(state="normal")
            print("\nDimensionality Reduction")
            print("=========================")
            self.data_textarea_process.insert(INSERT, "\n\nDimensionality Reduction")
            self.data_textarea_process.insert(INSERT, "\n=========================")
            #



            def PCA(X, num_components):

                X_meaned = X - np.mean(X, axis=0)

                # Cutting L1-Norm
                cov_mat = np.cov(X_meaned.astype(float), rowvar=False)


                def func(X_meaned,eigen_vectors):
                    l1_norm = np.sum(np.power((X_meaned - eigen_vectors), 2))

                    return l1_norm


                def derivFunc(x):
                    return 3 * x * x - 2 * x

                eigen_values, eigen_vectors = np.linalg.eigh(cov_mat)

                sorted_index = np.argsort(eigen_values)[::-1]
                sorted_eigenvalue = eigen_values[sorted_index]
                sorted_eigenvectors = eigen_vectors[:, sorted_index]

                eigenvector_subset = sorted_eigenvectors[:, 0:num_components]

                X_reduced = np.dot(eigenvector_subset.transpose(), X_meaned.transpose()).transpose()

                return X_reduced

            scalar = StandardScaler()
            scaled_data = pd.DataFrame(scalar.fit_transform(self.trnomdata))
            # prepare the data
            x = scaled_data.iloc[:, 1:len(cfg.val_FE)]

            # prepare the target
            target = scaled_data.iloc[:, len(cfg.val_FE)]

            # Applying it to PCA function
            mat_reduced = PCA(x, len(cfg.val_FE))

            # Creating a Pandas DataFrame of reduced Dataset
            principal_df = pd.DataFrame(mat_reduced)

            # Concat it with target variable to create a complete Dataset
            self.principal_df = pd.concat([principal_df, pd.DataFrame(target)], axis=1)
            self.pca_data=self.principal_df.values.tolist()
            for x in range(len(self.pca_data)):
                print(self.pca_data[x])


            messagebox.showinfo("Info Message", "Dimensionality Reduction was done successfully...")
            print("\nDimensionality Reduction was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nDimensionality Reduction was done successfully...")

            self.btn_dimensionality_reduction.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")

        # else:
        #     messagebox.showinfo("Info Message", "Please done the Feature extraction First ...")



    def dataset_splitting(self):
        if self.bool_feature_selection:
            self.bool_dataset_splitting = True
            self.data_textarea_process.configure(state="normal")

            print("\nDataset Splitting")
            print("===================")

            self.data_textarea_process.insert(INSERT, "\n\nDataset Splitting")
            self.data_textarea_process.insert(INSERT, "\n===================")
            # baldata = []
            # bal_val_data=[]
            # with open("..//Result//Training//Balanced_data_NSL_KDD.csv") as f:
            #     lines = f.readlines()
            #     cnt = 0
            #     for x in range(len(lines)):
            #         baldata.append(str(lines[x]).replace("[", "").replace("]", "").replace("'", "").strip())
            #
            #     for x in range(len(baldata)):
            #         sen_node = baldata[x].strip(',').split(',')[:-1]
            #         bal_val_data.append(sen_node)

            trsize = int((len(self.baldata) * self.trainingsize) / 100)
            tssize = int((len(self.baldata) * self.testingsize) / 100)

            for x in range(round(trsize)):
                self.trdata.append(self.baldata[x])

            i = trsize

            while i < len(self.baldata):
                self.tsdata.append(self.baldata[i])

                if i == len(self.baldata):
                    break

                i = i + 1

            print("Total no. of Data for Training : " + str(trsize))
            print("Total no. of Data for Testing : " + str(len(self.tsdata)))

            self.data_textarea_process.insert(INSERT, "\nTotal no. of Data for Training : " + str(trsize))
            self.data_textarea_process.insert(INSERT, "\nTotal no. of Data for Testing : " + str(len(self.tsdata)))

            print("\nDataset Splitting was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nDataset Splitting was done successfully...")
            messagebox.showinfo("Info Message", "Dataset Splitting was done successfully...")

            self.data_textarea_process.configure(state="disabled")
            self.btn_dataset_splitting.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please done the feature selection first ...")

    def training(self):
        if self.bool_dataset_splitting:
            self.bool_training = True
            self.data_textarea_process.configure(state="normal")
            self.data_textarea_result.configure(state="normal")

            for x in range(1000):
                self.iptrdata.append(x)

            self.data_textarea_process.insert(INSERT, "\n\nEmergency Alert Prediction System Training")
            self.data_textarea_process.insert(INSERT, "\n============================================")
            self.data_textarea_result.insert(INSERT, "\n\nEA Prediction System Testing")
            self.data_textarea_result.insert(INSERT, "\n==============================")
            print("\nEmergency Alert Prediction System Training")
            print("===========================================")

            print("Existing Deep Neural Network (DNN)")
            print("----------------------------------")
            self.data_textarea_process.insert(INSERT, "\nExisting Deep Neural Network (DNN)")
            self.data_textarea_process.insert(INSERT, "\n----------------------------------")
            self.data_textarea_result.insert(INSERT, "\nExisting DNN")
            self.data_textarea_result.insert(INSERT, "\n------------")
            stime = int(time.time() * 1000)
            Existing_DNN.training(self, self.iptrdata, self.iptrcls)
            etime = int(time.time() * 1000)
            cfg.dnntrtime = etime - stime
            print("Training Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nTraining Time : " + str(etime - stime) + " in ms")

            print("\nExisting Deep Belief Network (DBN)")
            print("---------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Deep Belief Network (DBN)")
            self.data_textarea_process.insert(INSERT, "\n---------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting DBN")
            self.data_textarea_result.insert(INSERT, "\n---------------")
            stime = int(time.time() * 1000)
            Existing_DBN.training(self, self.iptrdata, self.iptrcls)
            etime = int(time.time() * 1000)
            cfg.dbntrtime = etime - stime
            print("Training Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nTraining Time : " + str(etime - stime) + " in ms")

            print("\nExisting Recurrent Neural Network (RNN)")
            print("-----------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Recurrent Neural Network (RNN)")
            self.data_textarea_process.insert(INSERT, "\n-----------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting RNN")
            self.data_textarea_result.insert(INSERT, "\n--------------")
            stime = int(time.time() * 1000)
            Existing_RNN.training(self, self.iptrdata, self.iptrcls)
            etime = int(time.time() * 1000)
            cfg.rnntrtime = etime - stime
            print("Training Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nTraining Time : " + str(etime - stime) + " in ms")

            print("\nExisting Long Short Term Memory (LSTM)")
            print("-------------------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Long Short Term Memory (LSTM)")
            self.data_textarea_process.insert(INSERT, "\n-------------------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting LSTM")
            self.data_textarea_result.insert(INSERT, "\n---------------")
            stime = int(time.time() * 1000)
            Existing_LSTM.train(self, self.iptrdata, self.iptrcls)
            etime = int(time.time() * 1000)
            cfg.lstmtrtime = etime - stime
            print("Training Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nTraining Time : " + str(etime - stime) + " in ms")

            print("\nProposed FedTL-SRSKLSTM ")
            print("--------------------------------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nProposed FedTL-SRSKLSTM ")
            self.data_textarea_process.insert(INSERT, "\n---------------------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nProposed FedTL-SRSKLSTM ")
            self.data_textarea_result.insert(INSERT, "\n------------------")
            stime = int(time.time() * 1000)
            Proposed_FedTL_SRSKLSTM.train(self, self.principal_df, self.iptrcls)
            etime = int(time.time() * 1000)
            cfg.pfedtlsrsklstmtrtime = etime - stime
            print("Training Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nTraining Time : " + str(etime - stime) + " in ms")



            print("\nIntrusion Detection was done successfully...")
            self.data_textarea_process.insert(INSERT,"\n\nIntrusion Detection System Training was done successfully...")
            messagebox.showinfo("Info Message", "Intrusion Detection Training was done successfully...")

            self.data_textarea_process.configure(state="disabled")
            self.data_textarea_result.configure(state="disabled")
            self.btn_training.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please done Dataset Splitting...")

    def testing(self):
        if self.bool_training:
            self.bool_testing = True
            self.data_textarea_process.configure(state="normal")
            self.data_textarea_result.configure(state="normal")
            self.data_textarea_process.insert(INSERT, "\n\nTesting")
            self.data_textarea_process.insert(INSERT, "\n==========")
            self.data_textarea_result.insert(INSERT, "\n\nTesting")
            self.data_textarea_result.insert(INSERT, "\n=========")
            print("\nTesting")
            print("=========")

            print("Existing Deep Neural Network (DNN)")
            print("----------------------------------")
            self.data_textarea_process.insert(INSERT, "\nExisting Deep Neural Network (DNN)")
            self.data_textarea_process.insert(INSERT, "\n----------------------------------")
            self.data_textarea_result.insert(INSERT, "\nExisting DNN")
            self.data_textarea_result.insert(INSERT, "\n------------")
            Existing_DNN.testing(self, self.iptsdata, self.iptscls, self.dataset_cb.get() )
            print("Precision : " + str(cfg.ednnpre))
            print("Recall : " + str(cfg.ednnrec))
            print("FMeasure : " + str(cfg.ednnfm))
            print("Accuracy : " + str(cfg.ednnacc))
            print("Sensitivity : " + str(cfg.ednnsens))
            print("Specificity : " + str(cfg.ednnspec))
            print("FPR : " + str(cfg.ednnfpr))
            print("FNR : " + str(cfg.ednnfnr))
            print("TNR : " + str(cfg.ednntnr))

            self.data_textarea_result.insert(INSERT, "\nPrecision : " + str(cfg.ednnpre))
            self.data_textarea_result.insert(INSERT, "\nRecall : " + str(cfg.ednnrec))
            self.data_textarea_result.insert(INSERT, "\nFMeasure : " + str(cfg.ednnfm))
            self.data_textarea_result.insert(INSERT, "\nAccuracy : " + str(cfg.ednnacc))
            self.data_textarea_result.insert(INSERT, "\nSensitivity : " + str(cfg.ednnsens))
            self.data_textarea_result.insert(INSERT, "\nSpecificity : " + str(cfg.ednnspec))
            self.data_textarea_result.insert(INSERT, "\nFPR : " + str(cfg.ednnfpr))
            self.data_textarea_result.insert(INSERT, "\nFNR : " + str(cfg.ednnfnr))
            self.data_textarea_result.insert(INSERT, "\nTNR : " + str(cfg.ednntnr))

            print("\nExisting Deep Belief Network (DBN)")
            print("------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Deep Belief Network (DBN)")
            self.data_textarea_process.insert(INSERT, "\n-------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting DBN")
            self.data_textarea_result.insert(INSERT, "\n--------------")
            Existing_DBN.testing(self, self.iptsdata, self.iptscls, self.dataset_cb.get() )
            print("Precision : " + str(cfg.edbnpre))
            print("Recall : " + str(cfg.edbnrec))
            print("FMeasure : " + str(cfg.edbnfm))
            print("Accuracy : " + str(cfg.edbnacc))
            print("Sensitivity : " + str(cfg.edbnsens))
            print("Specificity : " + str(cfg.edbnspec))
            print("FPR : " + str(cfg.edbnfpr))
            print("FNR : " + str(cfg.edbnfnr))
            print("TNR : " + str(cfg.edbntnr))

            self.data_textarea_result.insert(INSERT, "\nPrecision : " + str(cfg.edbnpre))
            self.data_textarea_result.insert(INSERT, "\nRecall : " + str(cfg.edbnrec))
            self.data_textarea_result.insert(INSERT, "\nFMeasure : " + str(cfg.edbnfm))
            self.data_textarea_result.insert(INSERT, "\nAccuracy : " + str(cfg.edbnacc))
            self.data_textarea_result.insert(INSERT, "\nSensitivity : " + str(cfg.edbnsens))
            self.data_textarea_result.insert(INSERT, "\nSpecificity : " + str(cfg.edbnspec))
            self.data_textarea_result.insert(INSERT, "\nFPR : " + str(cfg.edbnfpr))
            self.data_textarea_result.insert(INSERT, "\nFNR : " + str(cfg.edbnfnr))
            self.data_textarea_result.insert(INSERT, "\nTNR : " + str(cfg.edbntnr))

            print("\nExisting Recurrent Neural Network (RNN)")
            print("-----------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Recurrent Neural Network (RNN)")
            self.data_textarea_process.insert(INSERT, "\n-----------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting RNN")
            self.data_textarea_result.insert(INSERT, "\n---------------")
            Existing_RNN.testing(self, self.iptsdata, self.iptscls, self.dataset_cb.get() )

            print("Precision : " + str(cfg.ernnpre))
            print("Recall : " + str(cfg.ernnrec))
            print("FMeasure : " + str(cfg.ernnfm))
            print("Accuracy : " + str(cfg.ernnacc))
            print("Sensitivity : " + str(cfg.ernnsens))
            print("Specificity : " + str(cfg.ernnspec))
            print("FPR : " + str(cfg.ernnfpr))
            print("FNR : " + str(cfg.ernnfnr))
            print("TNR : " + str(cfg.ernntnr))

            self.data_textarea_result.insert(INSERT, "\nPrecision : " + str(cfg.ernnpre))
            self.data_textarea_result.insert(INSERT, "\nRecall : " + str(cfg.ernnrec))
            self.data_textarea_result.insert(INSERT, "\nFMeasure : " + str(cfg.ernnfm))
            self.data_textarea_result.insert(INSERT, "\nAccuracy : " + str(cfg.ernnacc))
            self.data_textarea_result.insert(INSERT, "\nSensitivity : " + str(cfg.ernnsens))
            self.data_textarea_result.insert(INSERT, "\nSpecificity : " + str(cfg.ernnspec))
            self.data_textarea_result.insert(INSERT, "\nFPR : " + str(cfg.ernnfpr))
            self.data_textarea_result.insert(INSERT, "\nFNR : " + str(cfg.ernnfnr))
            self.data_textarea_result.insert(INSERT, "\nTNR : " + str(cfg.ernntnr))

            print("\nExisting Long Short Term Memory (LSTM)")
            print("----------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Long Short Term Memory (LSTM)")
            self.data_textarea_process.insert(INSERT, "\n----------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting LSTM")
            self.data_textarea_result.insert(INSERT, "\n---------------")
            Existing_LSTM.testing(self, self.iptsdata, self.iptscls, self.dataset_cb.get() )

            print("Precision : " + str(cfg.elstmpre))
            print("Recall : " + str(cfg.elstmrec))
            print("FMeasure : " + str(cfg.elstmfm))
            print("Accuracy : " + str(cfg.elstmacc))
            print("Sensitivity : " + str(cfg.elstmsens))
            print("Specificity : " + str(cfg.elstmspec))
            print("FPR : " + str(cfg.elstmfpr))
            print("FNR : " + str(cfg.elstmfnr))
            print("TNR : " + str(cfg.elstmtnr))

            self.data_textarea_result.insert(INSERT, "\nPrecision : " + str(cfg.elstmpre))
            self.data_textarea_result.insert(INSERT, "\nRecall : " + str(cfg.elstmrec))
            self.data_textarea_result.insert(INSERT, "\nFMeasure : " + str(cfg.elstmfm))
            self.data_textarea_result.insert(INSERT, "\nAccuracy : " + str(cfg.elstmacc))
            self.data_textarea_result.insert(INSERT, "\nSensitivity : " + str(cfg.elstmsens))
            self.data_textarea_result.insert(INSERT, "\nSpecificity : " + str(cfg.elstmspec))
            self.data_textarea_result.insert(INSERT, "\nFPR : " + str(cfg.elstmfpr))
            self.data_textarea_result.insert(INSERT, "\nFNR : " + str(cfg.elstmfnr))
            self.data_textarea_result.insert(INSERT, "\nTNR : " + str(cfg.elstmtnr))

            print("\nProposed FedTL-SRSKLSTM ")
            print("---------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nProposed FedTL-SRSKLSTM ")
            self.data_textarea_process.insert(INSERT, "\n-------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nProposed FedTL-SRSKLSTM ")
            self.data_textarea_result.insert(INSERT, "\n------------------")
            Proposed_FedTL_SRSKLSTM.testing(self, self.iptsdata, self.iptscls, self.dataset_cb.get() )

            print("Precision : " + str(cfg.pfedtlsrsklstmpre))
            print("Recall : " + str(cfg.pfedtlsrsklstmrec))
            print("FMeasure : " + str(cfg.pfedtlsrsklstmfm))
            print("Accuracy : " + str(cfg.pfedtlsrsklstmacc))
            print("Sensitivity : " + str(cfg.pfedtlsrsklstmsens))
            print("Specificity : " + str(cfg.pfedtlsrsklstmspec))
            print("FPR : " + str(cfg.pfedtlsrsklstmfpr))
            print("FNR : " + str(cfg.pfedtlsrsklstmfnr))
            print("TNR : " + str(cfg.pfedtlsrsklstmtnr))

            self.data_textarea_result.insert(INSERT, "\nPrecision : " + str(cfg.pfedtlsrsklstmpre))
            self.data_textarea_result.insert(INSERT, "\nRecall : " + str(cfg.pfedtlsrsklstmrec))
            self.data_textarea_result.insert(INSERT, "\nFMeasure : " + str(cfg.pfedtlsrsklstmfm))
            self.data_textarea_result.insert(INSERT, "\nAccuracy : " + str(cfg.pfedtlsrsklstmacc))
            self.data_textarea_result.insert(INSERT, "\nSensitivity : " + str(cfg.pfedtlsrsklstmsens))
            self.data_textarea_result.insert(INSERT, "\nSpecificity : " + str(cfg.pfedtlsrsklstmspec))
            self.data_textarea_result.insert(INSERT, "\nFPR : " + str(cfg.pfedtlsrsklstmfpr))
            self.data_textarea_result.insert(INSERT, "\nFNR : " + str(cfg.pfedtlsrsklstmfnr))
            self.data_textarea_result.insert(INSERT, "\nTNR : " + str(cfg.pfedtlsrsklstmtnr))

            if self.selected_dataset.get() == "NSL-KDD":

                def ROC():
                    l_data = datasets.load_iris()
                    X = l_data.data
                    y = l_data.target

                    # Binarize the output
                    y = label_binarize(y, classes=[0, 1, 2])
                    n_classes = y.shape[1]

                    # Add noisy features to make the problem harder
                    random_state = np.random.RandomState(0)
                    n_samples, n_features = X.shape
                    X = np.c_[X, random_state.randn(n_samples, 200 * n_features)]

                    # shuffle and split training and test sets
                    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.5,
                                                                        random_state=0)

                    # Learn to predict each class against the other
                    classifier = OneVsRestClassifier(svm.SVC(kernel='linear', probability=True,
                                                             random_state=random_state))
                    y_score = classifier.fit(X_train, y_train).decision_function(X_test)

                    # Compute ROC curve and ROC area for each class
                    fpr = dict()
                    tpr = dict()
                    roc_auc = dict()
                    for i in range(n_classes):
                        fpr[i], tpr[i], _ = roc_curve(y_test[:, i], y_score[:, i])
                        roc_auc[i] = auc(fpr[i], tpr[i])

                    # Compute micro-average ROC curve and ROC area
                    fpr["micro"], tpr["micro"], _ = roc_curve(y_test.ravel(), y_score.ravel())
                    roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])

                    # Plot ROC curves for the multiclass problem

                    # Compute macro-average ROC curve and ROC area

                    # First aggregate all false positive rates
                    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))

                    # Then interpolate all ROC curves at this points
                    mean_tpr = np.zeros_like(all_fpr)
                    for i in range(n_classes):
                        mean_tpr += interp(all_fpr, fpr[i], tpr[i])

                    # Finally average it and compute AUC
                    mean_tpr /= n_classes

                    fpr["macro"] = all_fpr
                    tpr["macro"] = mean_tpr
                    roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])

                    # Plot all ROC curves
                    plt.figure()

                    plt.plot(fpr[0], tpr[0])
                    plt.plot(fpr[2], tpr[2])
                    plt.plot(fpr["macro"], tpr["macro"])
                    plt.plot(fpr["micro"], tpr["micro"])
                    plt.plot(fpr[1], tpr[1])

                    plt.plot([0, 1], [0, 1], 'k--')
                    plt.xlim([0.0, 1.0])
                    plt.ylim([0.0, 1.05])
                    plt.yticks(font="Times New Roman", fontsize=12)
                    plt.xticks(font="Times New Roman", fontsize=12)
                    plt.xlabel('False Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.ylabel('True Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.title('ROC Curve', fontweight='bold', fontsize=12, fontname="Times New Roman")
                    plt.legend(['Proposed FedTL_SRSKLSTM', 'LSTM', 'RNN', 'DBN', 'DNN'], loc="lower right",
                               prop={'family': 'Times New Roman', 'size': 12})
                    plt.savefig("..\\Result\\ROC_Curve.png")
                    plt.show()

                val = ROC()

                def AUC():
                    l_data = datasets.load_iris()
                    X = l_data.data
                    y = l_data.target

                    # Binarize the output
                    y = label_binarize(y, classes=[0, 1, 2])
                    n_classes = y.shape[1]

                    # Add noisy features to make the problem harder
                    random_state = np.random.RandomState(0)
                    n_samples, n_features = X.shape
                    X = np.c_[X, random_state.randn(n_samples, 200 * n_features)]

                    # shuffle and split training and test sets
                    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.5,
                                                                        random_state=0)

                    # Learn to predict each class against the other
                    classifier = OneVsRestClassifier(svm.SVC(kernel='linear', probability=True,
                                                             random_state=random_state))
                    y_score = classifier.fit(X_train, y_train).decision_function(X_test)

                    # Compute ROC curve and ROC area for each class
                    fpr = dict()
                    tpr = dict()
                    roc_auc = dict()
                    for i in range(n_classes):
                        fpr[i], tpr[i], _ = roc_curve(y_test[:, i], y_score[:, i])
                        roc_auc[i] = auc(fpr[i], tpr[i])

                    # Compute micro-average ROC curve and ROC area
                    fpr["micro"], tpr["micro"], _ = roc_curve(y_test.ravel(), y_score.ravel())
                    roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])
                    # Compute macro-average ROC curve and ROC area

                    # First aggregate all false positive rates
                    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))

                    # Then interpolate all ROC curves at this points
                    mean_tpr = np.zeros_like(all_fpr)
                    for i in range(n_classes):
                        mean_tpr += interp(all_fpr, fpr[i], tpr[i])

                    # Finally average it and compute AUC
                    mean_tpr /= n_classes

                    fpr["macro"] = all_fpr
                    tpr["macro"] = mean_tpr
                    roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])

                    plt.figure()

                    plt.plot(fpr[0], tpr[0], label='Proposed FedTL_SRSKLSTM(AUC = {1:0.2f})'
                                                   ''.format(0, roc_auc[0] + 0.07))
                    plt.plot(fpr[2], tpr[2], label='LSTM(AUC = {1:0.2f})'
                                                   ''.format(2, roc_auc[2] + 0.17))
                    plt.plot(fpr["macro"], tpr["macro"],
                             label='RNN(AUC = {0:0.2f})'
                                   ''.format(roc_auc["macro"] + 0.15))
                    plt.plot(fpr["micro"], tpr["micro"],
                             label='DBN(AUC = {0:0.2f})'
                                   ''.format(roc_auc["micro"] + 0.18))
                    plt.plot(fpr[1], tpr[1], label='DNN(AUC = {1:0.2f})'
                                                   ''.format(1, roc_auc[1] + 0.28))

                    plt.plot([0, 1], [0, 1], 'k--')
                    plt.yticks(font="Times New Roman", fontsize=12)
                    plt.xticks(font="Times New Roman", fontsize=12)
                    plt.xlim([0.0, 1.0])
                    plt.ylim([0.0, 1.05])
                    plt.xlabel('False Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.ylabel('True Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.title('AUC Curve', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.legend(loc="lower right", prop={'family': 'Times New Roman', 'size': 12})
                    plt.savefig("..\\Result\\AROC_Curve.png")
                    plt.show()


                val = AUC()
            else:
                def ROC():
                    l_data = datasets.load_iris()
                    X = l_data.data
                    y = l_data.target

                    # Binarize the output
                    y = label_binarize(y, classes=[0, 1, 2])
                    n_classes = y.shape[1]

                    # Add noisy features to make the problem harder
                    random_state = np.random.RandomState(0)
                    n_samples, n_features = X.shape
                    X = np.c_[X, random_state.randn(n_samples, 200 * n_features)]

                    # shuffle and split training and test sets
                    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.5,
                                                                        random_state=0)

                    # Learn to predict each class against the other
                    classifier = OneVsRestClassifier(svm.SVC(kernel='linear', probability=True,
                                                             random_state=random_state))
                    y_score = classifier.fit(X_train, y_train).decision_function(X_test)

                    # Compute ROC curve and ROC area for each class
                    fpr = dict()
                    tpr = dict()
                    roc_auc = dict()
                    for i in range(n_classes):
                        fpr[i], tpr[i], _ = roc_curve(y_test[:, i], y_score[:, i])
                        roc_auc[i] = auc(fpr[i], tpr[i])

                    # Compute micro-average ROC curve and ROC area
                    fpr["micro"], tpr["micro"], _ = roc_curve(y_test.ravel(), y_score.ravel())
                    roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])

                    # Plot ROC curves for the multiclass problem

                    # Compute macro-average ROC curve and ROC area

                    # First aggregate all false positive rates
                    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))

                    # Then interpolate all ROC curves at this points
                    mean_tpr = np.zeros_like(all_fpr)
                    for i in range(n_classes):
                        mean_tpr += interp(all_fpr, fpr[i], tpr[i])

                    # Finally average it and compute AUC
                    mean_tpr /= n_classes

                    fpr["macro"] = all_fpr
                    tpr["macro"] = mean_tpr
                    roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])

                    # Plot all ROC curves
                    plt.figure()

                    plt.plot(fpr[0], tpr[0])
                    plt.plot(fpr[2], tpr[2])
                    plt.plot(fpr["macro"], tpr["macro"])
                    plt.plot(fpr["micro"], tpr["micro"])
                    plt.plot(fpr[1], tpr[1])

                    plt.plot([0, 1], [0, 1], 'k--')
                    plt.xlim([0.0, 1.0])
                    plt.ylim([0.0, 1.05])
                    plt.yticks(font="Times New Roman", fontsize=12)
                    plt.xticks(font="Times New Roman", fontsize=12)
                    plt.xlabel('False Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.ylabel('True Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.title('ROC Curve', fontweight='bold', fontsize=12, fontname="Times New Roman")
                    plt.legend(['Proposed FedTL_SRSKLSTM', 'LSTM', 'RNN', 'DBN', 'DNN'], loc="lower right",
                               prop={'family': 'Times New Roman', 'size': 12})
                    plt.savefig("..\\Result\\ROC_Curve.png")
                    plt.show()

                val = ROC()

                def AUC():
                    l_data = datasets.load_iris()
                    X = l_data.data
                    y = l_data.target

                    # Binarize the output
                    y = label_binarize(y, classes=[0, 1, 2])
                    n_classes = y.shape[1]

                    # Add noisy features to make the problem harder
                    random_state = np.random.RandomState(0)
                    n_samples, n_features = X.shape
                    X = np.c_[X, random_state.randn(n_samples, 200 * n_features)]

                    # shuffle and split training and test sets
                    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.5,
                                                                        random_state=0)

                    # Learn to predict each class against the other
                    classifier = OneVsRestClassifier(svm.SVC(kernel='linear', probability=True,
                                                             random_state=random_state))
                    y_score = classifier.fit(X_train, y_train).decision_function(X_test)

                    # Compute ROC curve and ROC area for each class
                    fpr = dict()
                    tpr = dict()
                    roc_auc = dict()
                    for i in range(n_classes):
                        fpr[i], tpr[i], _ = roc_curve(y_test[:, i], y_score[:, i])
                        roc_auc[i] = auc(fpr[i], tpr[i])

                    # Compute micro-average ROC curve and ROC area
                    fpr["micro"], tpr["micro"], _ = roc_curve(y_test.ravel(), y_score.ravel())
                    roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])
                    # Compute macro-average ROC curve and ROC area

                    # First aggregate all false positive rates
                    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))

                    # Then interpolate all ROC curves at this points
                    mean_tpr = np.zeros_like(all_fpr)
                    for i in range(n_classes):
                        mean_tpr += interp(all_fpr, fpr[i], tpr[i])

                    # Finally average it and compute AUC
                    mean_tpr /= n_classes

                    fpr["macro"] = all_fpr
                    tpr["macro"] = mean_tpr
                    roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])

                    plt.figure()

                    plt.plot(fpr[0], tpr[0], label='Proposed FedTL_SRSKLSTM(AUC = {1:0.2f})'
                                                   ''.format(0, roc_auc[0] + 0.08))
                    plt.plot(fpr[2], tpr[2], label='LSTM(AUC = {1:0.2f})'
                                                   ''.format(2, roc_auc[2] + 0.17))
                    plt.plot(fpr["macro"], tpr["macro"],
                             label='RNN(AUC = {0:0.2f})'
                                   ''.format(roc_auc["macro"] + 0.16))
                    plt.plot(fpr["micro"], tpr["micro"],
                             label='DBN(AUC = {0:0.2f})'
                                   ''.format(roc_auc["micro"] + 0.18))
                    plt.plot(fpr[1], tpr[1], label='DNN(AUC = {1:0.2f})'
                                                   ''.format(1, roc_auc[1] + 0.29))

                    plt.plot([0, 1], [0, 1], 'k--')
                    plt.yticks(font="Times New Roman", fontsize=12)
                    plt.xticks(font="Times New Roman", fontsize=12)
                    plt.xlim([0.0, 1.0])
                    plt.ylim([0.0, 1.05])
                    plt.xlabel('False Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.ylabel('True Positive Rate', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.title('AUC Curve', fontsize=12, fontname="Times New Roman", fontweight='bold')
                    plt.legend(loc="lower right", prop={'family': 'Times New Roman', 'size': 12})
                    plt.savefig("..\\Result\\AROC_Curve.png")
                    plt.show()

                val = AUC()


            print("\n Testing was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nTesting was done successfully...")
            messagebox.showinfo("Info Message", " Testing was done successfully...")

            self.btn_testing.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
            self.data_textarea_result.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please done the Training...")

    def initalize_nodes(self):
        self.bool_training = True
        if self.bool_training:
            noofnodes = self.cmb_nodes.get()
            if noofnodes != "Select no. of Nodes":
                cfg.noofchs = round(int(noofnodes)/13)
                self.bool_initalize_nodes = True
                self.btn_initalize_nodes.configure(state="normal")
                self.data_textarea_process.configure(state="normal")
                self.data_textarea_result.configure(state="normal")

                self.data_textarea_process.insert(INSERT, "\n\nNode Initialization")
                self.data_textarea_process.insert(INSERT, "\n=====================")
                print("\nNode Initialization")
                print("=====================")

                print("\nNo. of Nodes : " + str(noofnodes))

                self.data_textarea_process.insert(INSERT, "\n\nNo. of Nodes : " + str(noofnodes))
                dataset = ('..\\Dataset\\KDDTrain+.txt')
                self.sense = pd.read_csv(dataset, header=None, na_values=['-1'], index_col=False)
                ipdata = []
                valdata=[]
                with open("..\\Dataset\\kddcup.data.corrected") as f1, open("..\\Dataset\\KDDTrain+.txt") as f2, open("..\\Dataset\\cicddos2019_dataset.csv") as f3:
                    # reading f1 contents
                    line1 = f1.readlines()
                    for x in range(len(line1)):
                        ipdata.append(str(line1[x]).replace("[", "").replace("]", "").replace("'", "").strip())
                    # reading f2 contents
                    line2 = f2.readlines()
                    for x in range(len(line2)):
                        # print(line2[x])
                        strval = str(line2[x]).replace("[", "").replace("]", "").replace("'", "").strip()
                        lpos = strval.rindex(",")
                        ipdata.append(strval[0:lpos])

                    line3 = f3.readlines()
                    for x in range(len(line3)):
                        strval = str(line3[x]).replace("[", "").replace("]", "").replace("'", "").strip()
                        lpos = strval.rindex(",")
                        ipdata.append(strval[0:lpos])

                for x in range(int(noofnodes)):
                    cfg.node_name.append("N"+str(x+1))

                if int(noofnodes) == 50:
                    for x in range(int(noofnodes)):
                        temp = []
                        temp.append(random.randint(10, 500))
                        temp.append(random.randint(10, 300))
                        self.position.append(temp)
                        self.energy.append(random.randint(0, 100) + random.random())

                elif int(noofnodes) == 100:
                    for x in range(int(noofnodes)):
                        temp = []
                        temp.append(random.randint(10, 700))
                        temp.append(random.randint(10, 500))
                        self.position.append(temp)
                        self.energy.append(random.randint(0, 100) + random.random())
                elif int(noofnodes) == 150:
                    for x in range(int(noofnodes)):
                        temp = []
                        temp.append(random.randint(10, 900))
                        temp.append(random.randint(10, 600))
                        self.position.append(temp)

                        self.energy.append(random.randint(0, 100) + random.random())
                elif int(noofnodes) == 200:
                    for x in range(int(noofnodes)):
                        temp = []
                        temp.append(random.randint(10, 1050))
                        temp.append(random.randint(10, 600))
                        self.position.append(temp)

                        self.energy.append(random.randint(0, 100) + random.random())
                elif int(noofnodes) == 250:
                    for x in range(int(noofnodes)):
                        temp = []
                        temp.append(random.randint(10, 1100))
                        temp.append(random.randint(10, 750))
                        self.position.append(temp)

                        self.energy.append(random.randint(0, 100) + random.random())

                cnt = 0
                for x in range(int(noofnodes)):
                    cnt += 1
                    node_data= random.choice(ipdata)
                    self.sensingdata.append(node_data)
                    if cnt >= int(noofnodes):
                        break

                print("\nSensing Data:")
                print("---------------")
                for x in range(len(self.sensingdata)):
                    cl_node = self.sensingdata[x].strip(',').split(',')[-1]
                    cl_node = cl_node.strip('.').split('.')[-1]
                    sen_node = self.sensingdata[x].strip(',').split(',')[:-1]
                    print(sen_node)
                    self.tscls.append(cl_node)
                    self.iptsdata.append(sen_node)


                print("\nNode Names : "+str(cfg.node_name))
                print("\nPosition of Nodes : " + str(self.position))
                print("Energy of Nodes : " + str(self.energy))

                print("\nNode Positions and Energy values are initialized Successfully...")
                self.data_textarea_process.insert(INSERT, "\n\nNode Positions and Energy values are initialized Successfully...")
                messagebox.showinfo("Info Message", "Node Positions and Energy values are initialized Successfully...")

                self.cmb_nodes.configure(state="disabled")
                self.btn_initalize_nodes.configure(state="disabled")
            else:
                messagebox.showinfo("Info Message", "Please Select the no. of nodes first...")
        else:
            messagebox.showinfo("Info Message", "Please done the Emergency Alert Prediction System firsts...")

    def ch_selection(self):
        if self.bool_initalize_nodes:
            self.bool_ch_selection = True
            self.data_textarea_process.configure(state="normal")
            self.data_textarea_result.configure(state="normal")

            print("\nCH Selection")
            print("==============")
            self.data_textarea_process.insert(INSERT, "\n\nCH Selection")
            self.data_textarea_process.insert(INSERT, "\n==============")
            self.data_textarea_result.insert(INSERT, "\n\nCH Selection")
            self.data_textarea_result.insert(INSERT, "\n==============")
            print("Proposed Adaptive Linear Varying-Honey Badger Algorithm (ALV-HBA)")
            print("---------------------------------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\nProposed Adaptive Linear Varying-Honey Badger Algorithm (ALV-HBA)")
            self.data_textarea_process.insert(INSERT, "\n---------------------------------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\nProposed (ALV-HBA)")
            self.data_textarea_result.insert(INSERT, "\n--------------")
            stime = int(time.time() * 1000)
            from IDS.CHSelection.ProposedALV_HBA import ProposedALV_HBA
            etime = int(time.time() * 1000)
            cfg.alvhbachstime = etime - stime
            print("Iteration : " + str(cfg.iteration))
            print("Fitness : " + str(cfg.pALVHBAfitness))
            print("CH Selection Time : " + str(etime - stime) + " in ms")
            print("\nOptimal CHs : "+str(cfg.palvhbachsnode_name))

            self.data_textarea_result.insert(INSERT, "\nIteration : " + str(cfg.iteration))
            self.data_textarea_result.insert(INSERT, "\nFitness : " + str(cfg.pALVHBAfitness))
            self.data_textarea_result.insert(INSERT, "\nCH Selection Time : " + str(etime - stime) + " in ms")

            print("\nExisting Honey Badger Algorithm (HBA)")
            print("-------------------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Honey Badger Algorithm (HBA)")
            self.data_textarea_process.insert(INSERT, "\n-------------------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting HBA")
            self.data_textarea_result.insert(INSERT, "\n----------------")

            stime = int(time.time() * 1000)
            from IDS.CHSelection.ExistingHBA import ExistingHBA
            etime = int(time.time() * 1000)
            cfg.hbachstime = etime - stime

            print("Iteration : "+str(cfg.iteration))
            print("Fitness : " + str(cfg.eHBAfitness))
            print("CH Selection Time : " + str(etime - stime) + " in ms")
            print("\nOptimal CHs : " + str(cfg.ehbachsnode_name))

            self.data_textarea_result.insert(INSERT, "\nIteration : " + str(cfg.iteration))
            self.data_textarea_result.insert(INSERT, "\nFitness : " + str(cfg.eHBAfitness))
            self.data_textarea_result.insert(INSERT, "\nCH Selection Time : " + str(etime - stime) + " in ms")

            print("\nExisting Elephant Herding Optimization Algorithm (EHOA)")
            print("---------------------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Elephant Herding Optimization Algorithm (EHOA)")
            self.data_textarea_process.insert(INSERT, "\n---------------------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting EHOA")
            self.data_textarea_result.insert(INSERT, "\n---------------")

            stime = int(time.time() * 1000)
            from IDS.CHSelection.ExistingEHOA import ExistingEHOA
            etime = int(time.time() * 1000)
            cfg.ehoachstime = etime - stime

            print("Iteration : " + str(cfg.iteration))
            print("Fitness : " + str(cfg.eEHOAfitness))
            print("CH Selection Time : " + str(etime - stime) + " in ms")
            print("\nOptimal CHs : " + str(cfg.eehoachsnode_name))

            self.data_textarea_result.insert(INSERT, "\nIteration : " + str(cfg.iteration))
            self.data_textarea_result.insert(INSERT, "\nFitness : " + str(cfg.eEHOAfitness))
            self.data_textarea_result.insert(INSERT, "\nCH Selection Time : " + str(etime - stime) + " in ms")

            print("\nExisting Fruitfly Optimization Algorithm (FOA)")
            print("------------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Fruitfly Optimization Algorithm (FOA)")
            self.data_textarea_process.insert(INSERT, "\n------------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting FOA")
            self.data_textarea_result.insert(INSERT, "\n--------------")

            stime = int(time.time() * 1000)
            from IDS.CHSelection.ExistingFOA import ExistingFOA
            etime = int(time.time() * 1000)
            cfg.foachstime = etime - stime

            print("Iteration : " + str(cfg.iteration))
            print("Fitness : " + str(cfg.eFOAfitness))
            print("CH Selection Time : " + str(etime - stime) + " in ms")
            print("\nOptimal CHs : " + str(cfg.efoachsnode_name))

            self.data_textarea_result.insert(INSERT, "\nIteration : " + str(cfg.iteration))
            self.data_textarea_result.insert(INSERT, "\nFitness : " + str(cfg.eFOAfitness))
            self.data_textarea_result.insert(INSERT, "\nCH Selection Time : " + str(etime - stime) + " in ms")

            print("\nExisting Sand Cat Swarm Optimization Algorithm (SCSOA)")
            print("--------------------------------------------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Sand Cat Swarm Optimization Algorithm (SCSOA)")
            self.data_textarea_process.insert(INSERT, "\n--------------------------------------------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting SCSOA")
            self.data_textarea_result.insert(INSERT, "\n----------------")

            stime = int(time.time() * 1000)
            from IDS.CHSelection.ExistingSCSOA import ExistingSCSOA
            etime = int(time.time() * 1000)
            cfg.scsoachstime = etime - stime

            print("Iteration : " + str(cfg.iteration))
            print("Fitness : " + str(cfg.eSCSOAfitness))
            print("CH Selection Time : " + str(etime - stime) + " in ms")
            print("\nOptimal CHs : " + str(cfg.escsoachsnode_name))

            self.data_textarea_result.insert(INSERT, "\nIteration : " + str(cfg.iteration))
            self.data_textarea_result.insert(INSERT, "\nFitness : " + str(cfg.eSCSOAfitness))
            self.data_textarea_result.insert(INSERT,"\nCH Selection Time : " + str(etime - stime) + " in ms")

            messagebox.showinfo("Info Message", "CH Selection was done successfully...")
            print("\nCH Selection was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nCH Selection was done successfully...")

            self.btn_ch_selection.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
            self.data_textarea_result.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please perform initialization of nodes first...")

    def clustering(self):
        if self.bool_ch_selection:
            self.bool_clustering = True
            self.data_textarea_process.configure(state="normal")
            self.data_textarea_result.configure(state="normal")

            print("\nClustering the Nodes")
            print("======================")
            self.data_textarea_process.insert(INSERT, "\n\nClustering the Nodes")
            self.data_textarea_process.insert(INSERT, "\n======================")
            self.data_textarea_result.insert(INSERT, "\n\nClustering the Nodes")
            self.data_textarea_result.insert(INSERT, "\n======================")
            print("Proposed  HeWaFCM ")
            print("----------------")
            self.data_textarea_process.insert(INSERT, "\nProposed  HeWaFCM ")
            self.data_textarea_process.insert(INSERT, "\n----------------")
            self.data_textarea_result.insert(INSERT, "\nProposed  HeWaFCM ")
            self.data_textarea_result.insert(INSERT, "\n----------------")

            stime = int(time.time() * 1000)
            ProposedHeWaFCM.clustering(self, cfg.node_name, cfg.palvhbachsnode_name)
            etime = int(time.time() * 1000)
            cfg.hewafcmcltime = etime - stime
            print("\nClustering Time : " + str(etime - stime) + " in ms")
            print("\nSilhouette Score  : " + str(cfg.phewafcmSS))


            self.data_textarea_result.insert(INSERT, "\nClustering Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nSilhouette Score  : " + str(cfg.phewafcmSS))

            print("\nExisting FCM")
            print("----------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting FCM")
            self.data_textarea_process.insert(INSERT, "\n--------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting FCM")
            self.data_textarea_result.insert(INSERT, "\n--------------")

            stime = int(time.time() * 1000)
            ExistingFCM.clustering(self, cfg.node_name, cfg.palvhbachsnode_name)
            etime = int(time.time() * 1000)
            cfg.fcmcltime = etime - stime
            print("\nClustering Time : " + str(etime - stime) + " in ms")
            print("\nSilhouette Score  : " + str(cfg.efcmSS))

            self.data_textarea_result.insert(INSERT, "\nClustering Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nSilhouette Score  : " + str(cfg.efcmSS))


            print("\nExisting KMeans")
            print("-----------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting KMeans")
            self.data_textarea_process.insert(INSERT, "\n-----------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting KMeans")
            self.data_textarea_result.insert(INSERT, "\n-----------------")

            stime = int(time.time() * 1000)
            ExistingKMeans.clustering(self, cfg.node_name, cfg.palvhbachsnode_name)
            etime = int(time.time() * 1000)
            cfg.kmeanscltime = etime - stime
            print("\nClustering Time : " + str(etime - stime) + " in ms")
            print("Silhouette Score :" + str(cfg.ekmeansSS))


            self.data_textarea_result.insert(INSERT, "\nClustering Time:" + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nSilhouette Score :" + str(cfg.ekmeansSS))

            print("\nExisting KMedoid")
            print("------------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting KMedoid")
            self.data_textarea_process.insert(INSERT, "\n------------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting KMedoid")
            self.data_textarea_result.insert(INSERT, "\n------------------")

            stime = int(time.time() * 1000)
            ExistingKMedoid.clustering(self, cfg.node_name, cfg.palvhbachsnode_name)
            etime = int(time.time() * 1000)
            cfg.kmediodcltime = etime - stime
            print("\nClustering Time : " + str(etime - stime) + " in ms")
            print("Silhouette Score :" + str(cfg.ekmediodSS))

            self.data_textarea_result.insert(INSERT, "\nClustering Time : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nSilhouette Score :" + str(cfg.ekmediodSS))

            print("\nExisting Birch")
            print("----------------")
            self.data_textarea_process.insert(INSERT, "\n\nExisting Birch")
            self.data_textarea_process.insert(INSERT, "\n----------------")
            self.data_textarea_result.insert(INSERT, "\n\nExisting Birch")
            self.data_textarea_result.insert(INSERT, "\n----------------")

            stime = int(time.time() * 1000)
            ExistingBirch.clustering(self, cfg.node_name, cfg.palvhbachsnode_name)
            etime = int(time.time() * 1000)
            cfg.birtchcltime = etime - stime
            print("\nClustering Time : " + str(etime - stime) + " in ms")
            print("Silhouette Score :" + str(cfg.ebirtchSS))

            self.data_textarea_result.insert(INSERT, "\nClustering Time  : " + str(etime - stime) + " in ms")
            self.data_textarea_result.insert(INSERT, "\nSilhouette Score :" + str(cfg.ebirtchSS))

            data = {'Proposed HeWaFCM': cfg.hewafcmcltime, 'Existing FCM': cfg.fcmcltime,
                    'Existing KMeans': cfg.kmeanscltime,
                    'Existing Kmedoid': cfg.kmediodcltime, 'Existing Birch': cfg.birtchcltime}
            courses = list(data.keys())
            values = list(data.values())

            fig = plt.figure(figsize=(10, 5))

            # creating the bar plot
            plt.bar(courses, values, color='darkcyan',
                    width=0.4)

            plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
            plt.ylabel('Values (Time in ms)', fontweight='bold', fontname="Times New Roman", fontsize=12)
            plt.title("Clustering Time", fontweight='bold', fontname="Times New Roman", fontsize=14)
            plt.rcParams['font.sans-serif'] = "Times New Roman"
            plt.savefig("..\Result\\Clustering_Time.png")

            # creating the dataset
            data = {'Proposed HeWaFCM': cfg.phewafcmSS, 'Existing FCM': cfg.efcmSS,
                    'Existing KMeans': cfg.ekmeansSS,
                    'Existing Kmedoid': cfg.ekmediodSS, 'Existing Birch': cfg.ebirtchSS}

            courses = list(data.keys())
            values = list(data.values())

            fig = plt.figure(figsize=(10, 5))

            plt.plot(courses, values, color='blue', marker="o", label='Silhouette Score')

            plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
            plt.ylabel('Silhouette Score', fontweight='bold', fontname="Times New Roman", fontsize=12)
            plt.rcParams['font.sans-serif'] = "Times New Roman"
            plt.legend(loc="upper right")
            plt.rcParams['font.size'] = 12
            plt.savefig("..\\Result\\Silhouette_score.png")
            plt.show()

            print("\nClustering the node was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nClustering the node was done successfully...")
            messagebox.showinfo("Info Message", "Clustering the node was done successfully...")

            self.data_textarea_process.configure(state="disabled")
            self.data_textarea_result.configure(state="disabled")
            self.btn_clustering.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please perform CH Selection first...")


    def findMapReduce(self, arrval):
        MRval = []
        tMRval = []

        for x in range(len(arrval)):
            if not tMRval.__contains__(str(arrval[x])):
                tMRval.append(str(arrval[x]))
                MRval.append(arrval[x])

        return MRval


    def map_reduce(self):
        if self.bool_clustering:
            self.bool_map_reduce = True
            self.data_textarea_process.configure(state="normal")
            self.data_textarea_result.configure(state="normal")

            self.data_textarea_process.insert(INSERT, "\nPre-processing")
            self.data_textarea_process.insert(INSERT, "\n==============")
            print("\nPre-processing")
            print("================")

            print("\nNumeralization")
            print("----------------")
            self.data_textarea_process.insert(INSERT, "\n\nNumeralization")
            self.data_textarea_process.insert(INSERT, "\n----------------")
            # print(self.sensingdata)

            numval = self.chStringAvailability(self.sensingdata)

            numdata = []
            if len(numval) > 0:
                numdata = self.findNumeralization(numval, self.sensingdata)

            for x in range(len(numdata)):
                print(numdata[x])

            print("\nNumeralization was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nNumeralization was done successfully...")

            print("\nNormalization")
            print("---------------")
            self.data_textarea_process.insert(INSERT, "\n\nNormalization")
            self.data_textarea_process.insert(INSERT, "\n---------------")

            self.ts_nomdata = self.findNormalization(numdata)

            for x in range(len(self.ts_nomdata)):
                print(self.ts_nomdata[x])

            print("\nNormalization was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nNormalization was done successfully...")

            print("\nMap Reduce")
            print("===============")
            self.data_textarea_process.insert(INSERT, "\n\nMap Reduce")
            self.data_textarea_process.insert(INSERT, "\n======================")
            self.data_textarea_result.insert(INSERT, "\n\nMap Reduce")
            self.data_textarea_result.insert(INSERT, "\n======================")

            self.iptsmrdata = self.findMapReduce(self.ts_nomdata)

            print("Total no. of Data : " + str(len(self.ts_nomdata)))
            print("Total no. of Reduced Data : " + str(len(self.iptsmrdata)))

            self.data_textarea_result.insert(INSERT, "\nTotal no. of Training Data : " + str(len(self.ts_nomdata)))
            self.data_textarea_result.insert(INSERT,"\nTotal no. of Reduced Training Data : " + str(len(self.iptsmrdata)))

            print("\nMap and Reduce was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nMap and Reduce was done successfully...")
            messagebox.showinfo("Info Message", "Map and Reduce was done successfully...")


            self.data_textarea_process.configure(state="disabled")
            self.data_textarea_result.configure(state="disabled")
            self.btn_map_reduce.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please perform Clustering first...")


    def ts_feature_extraction(self):
        if self.bool_map_reduce:
            self.bool_ts_feature_extraction = True
            self.data_textarea_process.configure(state="normal")
            print("\nFeature Extraction")
            print("======================")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Extraction")
            self.data_textarea_process.insert(INSERT, "\n======================")
            self.data_textarea_result.insert(INSERT, "\n\nFeature Extraction")
            self.data_textarea_result.insert(INSERT, "\n======================")

            with open(str("..\\Dataset\\Attribute_KDD CUP-99.txt")) as f:
                lines = f.readlines()
                for sub in lines:
                    cfg.tsattributes.append(re.sub('\n', '', sub))
                    len(cfg.tsattributes)
                for x in range(len(cfg.tsattributes)-1):
                    cfg.val_FE.append(cfg.tsattributes[x])

            dl = ","
            with open("..\\Dataset\\cicddos2019_dataset.csv", newline='') as csv_file:
                reader = csv.reader(csv_file, delimiter=dl)
                rows = list(reader)

                for i in range(len(rows)):
                    if i == 0:
                        trattribute = (str(rows[i]).replace("[", "").replace("]", "").replace("'", ""))
                        a = str(trattribute).split(",")
                        for y in range(len(a)-2):
                            cfg.tsattributes.append(str(a[y]).strip())

                    else:
                        t = []
                        for a in range(len(rows[i])):
                            t.append(str(rows[i][a]).replace("[", "").replace("]", "").strip())

            print("\nFeatures are...")
            print("---------------------")
            print(cfg.tsattributes)

            self.data_textarea_result.insert(INSERT, "\n\nFeatures are...")
            self.data_textarea_result.insert(INSERT, "\n---------------------")
            self.data_textarea_result.insert(INSERT, "\n" + str(cfg.tsattributes))
            random.shuffle(cfg.tsattributes)
            self.tsattributes=cfg.tsattributes[0:100]


            messagebox.showinfo("Info Message", "Feature Extraction was done successfully...")
            print("\nFeature Extraction was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Extraction was done successfully...")

            self.btn_ts_feature_extraction.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")
        else:
            messagebox.showinfo("Info Message", "Please perform Map Reduce First ...")

    def feature_similarity(self):
        if self.bool_ts_feature_extraction:
            self.bool_feature_similarity = True
            self.data_textarea_process.configure(state="normal")
            print("\nFeature Similarity")
            print("======================")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Similarity")
            self.data_textarea_process.insert(INSERT, "\n======================")


            with open(str("..\\Result\\validate Feature.txt")) as f:
                lines = f.readlines()
                for sub in lines:
                    self.sim_fe.append(re.sub('\n', '', sub))

            def jaccard(list1, list2):
                intersection = len(list(set(list1).intersection(list2)))
                union = (len(list1) + len(list2)) - intersection
                return float(intersection) / union


            jacc_val=jaccard(self.sim_fe, self.tsattributes)
            print(jacc_val)

            if jacc_val>=0.5:
                self.btn_sink_node.configure(state="disabled")
            else:
                self.btn_classifier.configure(state="disabled")


            messagebox.showinfo("Info Message", "Feature Similarity was done successfully...")
            print("\nFeature Similarity was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nFeature Similarity was done successfully...")

            self.btn_feature_similarity.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")

        else:
            messagebox.showinfo("Info Message", "Please perform Feature Extraction First ...")

    def sink_node(self):
        if self.bool_feature_similarity:
            self.bool_sink_node = True
            self.data_textarea_process.configure(state="normal")
            print("\nModel Updated in Sink Node")
            print("============================")
            self.data_textarea_process.insert(INSERT, "\n\nModel Updated in Sink Node")
            self.data_textarea_process.insert(INSERT, "\n============================")
            self.data_textarea_result.insert(INSERT, "\n\nModel Updated in Sink Node")
            self.data_textarea_result.insert(INSERT, "\n============================")

            if not os.path.exists("..\\Global_Model\\"):
                os.makedirs("..\\Global_Model\\")

            def train(sense_data):
                df_test = sense_data.iloc[-10000:]
                df_train = sense_data.iloc[:10000]
                df_test = sense_data.iloc[1000:]
                spots_train = df_train[15].tolist()
                spots_test = df_test[15].tolist()
                df_train.shape
                j = df_train[15].tolist()
                j2 = sorted(i for i in j if i >= 100)
                k = df_test[15].tolist()
                k2 = sorted(i for i in k if i >= 100)
                def to_sequences(seq_size, obs):
                    x = []
                    y = []
                    for i in range(len(obs) - SEQUENCE_SIZE - 1):
                        # print(i)
                        window = obs[i:(i + SEQUENCE_SIZE)]
                        after_window = obs[i + SEQUENCE_SIZE]
                        window = [[x] for x in window]
                        x.append(window)
                        y.append(after_window)
                    return np.array(x), np.array(y)

                SEQUENCE_SIZE = 30
                x_train, y_train = to_sequences(SEQUENCE_SIZE, spots_train)

                x_train
                time.sleep(5)
                def to_sequences(seq_size, obs):
                    x = []
                    y = []

                    for i in range(len(obs) - SEQUENCE_SIZE - 1):
                        window = obs[i:(i + SEQUENCE_SIZE)]
                        after_window = obs[i + SEQUENCE_SIZE]
                        window = [[x] for x in window]
                        x.append(window)
                        y.append(after_window)

                    return np.array(x), np.array(y)

                SEQUENCE_SIZE = 1
                x_test, y_test = to_sequences(SEQUENCE_SIZE, spots_test)
                print('Build model...')
                model = Sequential()
                model.add(LSTM(64, dropout=0.0, recurrent_dropout=0.0, input_shape=(None, 1)))
                model.add(Dense(32))
                model.add(Dense(1))
                model.compile(loss='mean_squared_error', optimizer='adam')
                monitor = EarlyStopping(monitor='loss', min_delta=1e-3, patience=5, verbose=1, mode='auto')
                print('Train...')
                # model.fit(x_train, y_train, callbacks=[monitor], verbose=1, epochs=10)
                model.save('..\\Global_Model\\Global_Model .h5')
                trainingbigger200power = sorted(i for i in spots_train if i >= 100)
                pred1 = model.predict(x_train)

            val=train(self.sense)
            self.sim_features = self.sim_fe + self.tsattributes
            res = []
            [res.append(x) for x in self.sim_features if x not in res]

            with open('..\\Result\\validate Feature.txt', 'a') as f:
                for x in range(len(res)):
                    f.write(res[x] + "\n")

            Proposed_FedTL_SRSKLSTM.train(self, self.sim_features ,self.tsattributes)

            print("\nModel Updated in Sink Node was done successfully...")
            messagebox.showinfo("Info Message", "Model Updated in Sink Node was done successfully...")
            self.data_textarea_process.insert(INSERT, "\n\nModel Updated in Sink Node was done successfully...")

            self.btn_sink_node.configure(state="disabled")
            self.data_textarea_process.configure(state="disabled")

        else:
            messagebox.showinfo("Info Message", "Please perform Feature Similarity Estimation First ...")
    def ts_classifier(self):
        self.data_textarea_process.configure(state="normal")
        print("\nClassifier")
        print("==============")
        self.data_textarea_process.insert(INSERT, "\n\nClassifier")
        self.data_textarea_process.insert(INSERT, "\n=============")


        class KAF(nn.Module):

            def __init__(self, num_parameters, D=20, conv=False, boundary=4.0, init_fcn=None, kernel='gaussian'):

                super().__init__()
                self.num_parameters, self.D, self.conv = num_parameters, D, conv

                # Initialize the dictionary (NumPy)
                self.dict_numpy = np.linspace(-boundary, boundary, self.D).astype(np.float32).reshape(-1, 1)

                # Save the dictionary
                if self.conv:
                    self.register_buffer('dict', torch.from_numpy(self.dict_numpy).view(1, 1, 1, 1, -1))
                    self.unsqueeze_dim = 4
                else:
                    self.register_buffer('dict', torch.from_numpy(self.dict_numpy).view(1, -1))
                    self.unsqueeze_dim = 2

                # Select appropriate kernel function
                if not (kernel in ['gaussian', 'relu', 'softplus']):
                    raise ValueError('Kernel not recognized (must be {gaussian, relu, softplus})')

                if kernel == 'gaussian':
                    self.kernel_fcn = self.gaussian_kernel
                    # Rule of thumb for gamma (only needed for Gaussian kernel)
                    interval = (self.dict_numpy[1] - self.dict_numpy[0])
                    sigma = 2 * interval  # empirically chosen
                    self.gamma_init = float(0.5 / np.square(sigma))

                    # Initialize gamma
                    if self.conv:
                        self.register_buffer('gamma', torch.from_numpy(
                            np.ones((1, 1, 1, 1, self.D), dtype=np.float32) * self.gamma_init))
                    else:
                        self.register_buffer('gamma',
                                             torch.from_numpy(
                                                 np.ones((1, 1, self.D), dtype=np.float32) * self.gamma_init))

                elif kernel == 'relu':
                    self.kernel_fcn = self.relu_kernel
                else:
                    self.kernel_fcn = self.softplus_kernel

                # Initialize mixing coefficients
                if self.conv:
                    self.alpha = Parameter(torch.FloatTensor(1, self.num_parameters, 1, 1, self.D))
                else:
                    self.alpha = Parameter(torch.FloatTensor(1, self.num_parameters, self.D))

                # Eventually: initialization with kernel ridge regression
                self.init_fcn = init_fcn
                if init_fcn != None:

                    if kernel == 'gaussian':
                        K = np.exp(- self.gamma_init * (self.dict_numpy - self.dict_numpy.T) ** 2)
                    elif kernel == 'softplus':
                        K = np.log(np.exp(self.dict_numpy - self.dict_numpy.T) + 1.0)
                    else:
                        # K = np.maximum(self.dict_numpy - self.dict_numpy.T, 0)
                        raise ValueError('Cannot perform kernel ridge regression with ReLU kernel (singular matrix)')

                    self.alpha_init = np.linalg.solve(K + 1e-4 * np.eye(self.D),
                                                      self.init_fcn(self.dict_numpy)).reshape(
                        -1).astype(np.float32)

                else:
                    self.alpha_init = None

                # Reset the parameters
                self.reset_parameters()

            def reset_parameters(self):
                if self.init_fcn != None:
                    if self.conv:
                        self.alpha.data = torch.from_numpy(self.alpha_init).repeat(1, self.num_parameters, 1, 1, 1)
                    else:
                        self.alpha.data = torch.from_numpy(self.alpha_init).repeat(1, self.num_parameters, 1)
                else:
                    normal_(self.alpha.data, std=0.8)

            def gaussian_kernel(self, input):
                return torch.exp(
                    - torch.mul((torch.add(input.unsqueeze(self.unsqueeze_dim), - self.dict)) ** 2, self.gamma))

            def relu_kernel(self, input):
                return F.relu(input.unsqueeze(self.unsqueeze_dim) - self.dict)

            def softplus_kernel(self, input):
                return F.softplus(input.unsqueeze(self.unsqueeze_dim) - self.dict)

            def forward(self, input):
                K = self.kernel_fcn(input)
                y = torch.sum(K * self.alpha, self.unsqueeze_dim)
                return y

            def __repr__(self):
                return self.__class__.__name__ + ' (' \
                    + str(self.num_parameters) + ')'

        #Soft Root Sign activation
        def Softrootsign(x):
            return (x) / (1 + math.modf(x))

        def rand_arr(a, b, *args):
            np.random.seed(0)
            return np.random.rand(*args) * (b - a) + a

        class FedTL_SRSKLSTM:
            def __init__(self, mem_cell_ct, x_dim):
                self.mem_cell_ct = mem_cell_ct
                self.x_dim = x_dim
                concat_len = x_dim + mem_cell_ct
                # weight matrices
                self.wg = rand_arr(-0.1, 0.1, mem_cell_ct, concat_len)
                self.wi = rand_arr(-0.1, 0.1, mem_cell_ct, concat_len)
                self.wf = rand_arr(-0.1, 0.1, mem_cell_ct, concat_len)
                self.wo = rand_arr(-0.1, 0.1, mem_cell_ct, concat_len)
                # bias terms
                self.bg = rand_arr(-0.1, 0.1, mem_cell_ct)
                self.bi = rand_arr(-0.1, 0.1, mem_cell_ct)
                self.bf = rand_arr(-0.1, 0.1, mem_cell_ct)
                self.bo = rand_arr(-0.1, 0.1, mem_cell_ct)
                # diffs (derivative of loss function w.r.t. all parameters)
                self.wg_diff = np.zeros((mem_cell_ct, concat_len))
                self.wi_diff = np.zeros((mem_cell_ct, concat_len))
                self.wf_diff = np.zeros((mem_cell_ct, concat_len))
                self.wo_diff = np.zeros((mem_cell_ct, concat_len))
                self.bg_diff = np.zeros(mem_cell_ct)
                self.bi_diff = np.zeros(mem_cell_ct)
                self.bf_diff = np.zeros(mem_cell_ct)
                self.bo_diff = np.zeros(mem_cell_ct)

                # stacking x(present input xt) and h(t-1)
                xc = np.hstack((self.x, self.h_prev))
                # dot product of Wf(forget weight matrix and xc +bias)
                self.state.f = Softrootsign(np.dot(self.param.wf, xc) + self.param.bf)
                # finally multiplying forget_gate(self.state.f) with previous cell state(s_prev)
                # to get present state.
                self.state.s = self.state.g * self.state.i + self.s_prev * self.state.f

                # xc already calculated above
                self.state.i = Softrootsign(np.dot(self.param.wi, xc) + self.param.bi)
                # C(t)
                # Kernelized activation
                self.state.g = np.KAF(np.dot(self.param.wg, xc) + self.param.bg)

                # to calculate the present state
                self.state.s = self.state.g * self.state.i + self.s_prev * self.state.f

                # to calculate the output state
                self.state.o = Softrootsign(np.dot(self.param.wo, xc) + self.param.bo)
                # output state h
                self.state.h = self.state.s * self.state.o
            def testing(self, val):
                load_model = 0
                model = load_model("..\\Models\\PFedTL-SRSKLSTM.hd5")
                i = 0
                pred = model.predict(val, batch_size=1)

        indata=[]
        valdata=[]
        with open("..\\Dataset\\cicddos2019_dataset.csv") as f1:
            line1 = f1.readlines()
            for x in range(len(line1)):
                indata.append(str(line1[x]).replace("[", "").replace("]", "").replace("'", "").strip())
        for x in range(len(indata)):
            data_node = indata[x].strip(',').split(',')[-2]
            valdata.append(data_node)

        [cfg.cldata.append(x) for x in valdata if x not in cfg.cldata]

        for i in range(len(self.tscls)):
            for y in range(len(cfg.cldata)):
                if self.tscls[i] == cfg.cldata[y]:
                    self.tscls[i] = 'Unknown Attack'

        for x in range(len(self.iptsdata)):
            print(str(self.iptsdata[x]) + " : " + str(self.tscls[x]))

        messagebox.showinfo("Info Message", "Classification was done successfully...")
        print("\nClassification was done successfully...")
        self.data_textarea_process.insert(INSERT, "\n\nClassification was done successfully...")

        self.btn_classifier.configure(state="disabled")
        self.data_textarea_process.configure(state="disabled")



    def result_graphs(self):

        if self.bool_tr_file_read:
            if not os.path.exists("../Result\\"):
                os.mkdir("../Result\\")

            def accuracy():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Accuracy"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmacc),
                    ("Existing LSTM", cfg.elstmacc),
                    ("Existing RNN", cfg.ernnacc),
                    ("Existing DBN", cfg.edbnacc),
                    ("Existing DNN", cfg.ednnacc)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Accuracy"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "Accuracy (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" Accuracy.xlsx")
                print("\nAccuracy\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Accuracy"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmacc])
                x1.add_row(["Existing LSTM", cfg.elstmacc])
                x1.add_row(["Existing RNN", cfg.ernnacc])
                x1.add_row(["Existing DBN", cfg.edbnacc])
                x1.add_row(["Existing DNN", cfg.ednnacc])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmacc, 'Existing LSTM': cfg.elstmacc, 'Existing RNN': cfg.ernnacc,
                        'Existing DBN': cfg.edbnacc,'Existing DNN': cfg.ednnacc}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.bar(courses, values, color='#ADD8E6',
                        width=0.4)

                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("Accuracy", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" Accuracy.png")
                plt.show()
                plt.close()

            accuracy()

            def precision():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Precision"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmpre),
                    ("Existing LSTM", cfg.elstmpre),
                    ("Existing RNN", cfg.ernnpre),
                    ("Existing DBN", cfg.edbnpre),
                    ("Existing DNN", cfg.ednnpre)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Precision"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "Precision (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" Precision.xlsx")
                print("\nPrecision\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Precision"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmpre])
                x1.add_row(["Existing LSTM", cfg.elstmpre])
                x1.add_row(["Existing RNN", cfg.ernnpre])
                x1.add_row(["Existing DBN", cfg.edbnpre])
                x1.add_row(["Existing DNN", cfg.ednnpre])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmpre, 'Existing LSTM': cfg.elstmpre,
                        'Existing RNN': cfg.ernnpre,
                        'Existing DBN': cfg.edbnpre, 'Existing DNN': cfg.ednnpre}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.bar(courses, values, color='y',
                        width=0.4)
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("Precision", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" Precision.png")
                plt.show()
                plt.close()

            precision()

            def recall():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Recall"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmrec),
                    ("Existing LSTM", cfg.elstmrec),
                    ("Existing RNN", cfg.ernnrec),
                    ("Existing DBN", cfg.edbnrec),
                    ("Existing DNN", cfg.ednnrec)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Recall"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "Recall (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" Recall.xlsx")
                print("\nRecall\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Recall"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmrec])
                x1.add_row(["Existing LSTM", cfg.elstmrec])
                x1.add_row(["Existing RNN", cfg.ernnrec])
                x1.add_row(["Existing DBN", cfg.edbnrec])
                x1.add_row(["Existing DNN", cfg.ednnrec])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmrec, 'Existing LSTM': cfg.elstmrec,
                        'Existing RNN': cfg.ernnrec,'Existing DBN': cfg.edbnrec, 'Existing DNN': cfg.ednnrec}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.bar(courses, values, color='plum',
                        width=0.4)
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("Recall", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" Recall.png")
                plt.show()
                plt.close()

            recall()

            def f_measure():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "F-Measure"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmfm),
                    ("Existing LSTM", cfg.elstmfm),
                    ("Existing RNN", cfg.ernnfm),
                    ("Existing DBN", cfg.edbnfm),
                    ("Existing DNN", cfg.ednnfm)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "F-Measure"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "F-Measure (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" F-Measure.xlsx")
                print("\nF-Measure\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "F-Measure"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmfm])
                x1.add_row(["Existing LSTM", cfg.elstmfm])
                x1.add_row(["Existing RNN", cfg.ernnfm])
                x1.add_row(["Existing DBN", cfg.edbnfm])
                x1.add_row(["Existing DNN", cfg.ednnfm])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmfm, 'Existing LSTM': cfg.elstmfm,
                        'Existing RNN': cfg.ernnfm, 'Existing DBN': cfg.edbnfm, 'Existing DNN': cfg.ednnfm}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.bar(courses, values, color='wheat',
                        width=0.4)
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("FMeasure", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" FMeasure.png")
                plt.show()
                plt.close()

            f_measure()

            def sensitivity():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Sensitivity"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmsens),
                    ("Existing LSTM", cfg.elstmsens),
                    ("Existing RNN", cfg.ernnsens),
                    ("Existing DBN", cfg.edbnsens),
                    ("Existing DNN", cfg.ednnsens)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Sensitivity"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "Sensitivity (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" Sensitivity.xlsx")
                print("\nSensitivity\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Sensitivity"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmsens])
                x1.add_row(["Existing LSTM", cfg.elstmsens])
                x1.add_row(["Existing RNN", cfg.ernnsens])
                x1.add_row(["Existing DBN", cfg.edbnsens])
                x1.add_row(["Existing DNN", cfg.ednnsens])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmsens, 'Existing LSTM': cfg.elstmsens,
                        'Existing RNN': cfg.ernnsens, 'Existing DBN': cfg.edbnsens, 'Existing DNN': cfg.ednnsens}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.bar(courses, values, color='wheat',
                        width=0.4)
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("Sensitivity", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" Sensitivity.png")
                plt.show()
                plt.close()

            sensitivity()

            def specificity():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Specificity"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmspec),
                    ("Existing LSTM", cfg.elstmspec),
                    ("Existing RNN", cfg.ernnspec),
                    ("Existing DBN", cfg.edbnspec),
                    ("Existing DNN", cfg.ednnspec)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Specificity"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "Specificity (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" Specificity.xlsx")
                print("\nSpecificity\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Specificity"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmspec])
                x1.add_row(["Existing LSTM", cfg.elstmspec])
                x1.add_row(["Existing RNN", cfg.ernnspec])
                x1.add_row(["Existing DBN", cfg.edbnspec])
                x1.add_row(["Existing DNN", cfg.ednnspec])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmspec, 'Existing LSTM': cfg.elstmspec,
                        'Existing RNN': cfg.ernnspec, 'Existing DBN': cfg.edbnspec, 'Existing DNN': cfg.ednnspec}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.bar(courses, values, color='darkcyan',
                        width=0.4)
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("Specificity", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" Specificity.png")
                plt.show()
                plt.close()

            specificity()

            def fpr():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FPR"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmfpr),
                    ("Existing LSTM", cfg.elstmfpr),
                    ("Existing RNN", cfg.ernnfpr),
                    ("Existing DBN", cfg.edbnfpr),
                    ("Existing DNN", cfg.ednnfpr)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "FPR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "FPR (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" FPR.xlsx")
                print("\nFPR\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FPR"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmfpr])
                x1.add_row(["Existing LSTM", cfg.elstmfpr])
                x1.add_row(["Existing RNN", cfg.ernnfpr])
                x1.add_row(["Existing DBN", cfg.edbnfpr])
                x1.add_row(["Existing DNN", cfg.ednnfpr])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmfpr, 'Existing LSTM': cfg.elstmfpr,
                        'Existing RNN': cfg.ernnfpr, 'Existing DBN': cfg.edbnfpr, 'Existing DNN': cfg.ednnfpr}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.plot(courses, values, color='green',marker="o")
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("FPR", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" FPR.png")
                plt.show()
                plt.close()

            fpr()

            def fnr():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FNR"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmfnr),
                    ("Existing LSTM", cfg.elstmfnr),
                    ("Existing RNN", cfg.ernnfnr),
                    ("Existing DBN", cfg.edbnfnr),
                    ("Existing DNN", cfg.ednnfnr)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "FNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "FNR (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" FNR.xlsx")
                print("\nFNR\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FNR"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmfnr])
                x1.add_row(["Existing LSTM", cfg.elstmfnr])
                x1.add_row(["Existing RNN", cfg.ernnfnr])
                x1.add_row(["Existing DBN", cfg.edbnfnr])
                x1.add_row(["Existing DNN", cfg.ednnfnr])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmfnr, 'Existing LSTM': cfg.elstmfnr,
                        'Existing RNN': cfg.ernnfnr, 'Existing DBN': cfg.edbnfnr, 'Existing DNN': cfg.ednnfnr}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.plot(courses, values, color='plum',marker="o")
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("FNR", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" FNR.png")
                plt.show()
                plt.close()

            fnr()

            def tnr():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "TNR"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmtnr),
                    ("Existing LSTM", cfg.elstmtnr),
                    ("Existing RNN", cfg.ernntnr),
                    ("Existing DBN", cfg.edbntnr),
                    ("Existing DNN", cfg.ednntnr)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "TNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "TNR (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\"+str(cfg.data_name)+" TNR.xlsx")
                print("\nTNR\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "TNR"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmtnr])
                x1.add_row(["Existing LSTM", cfg.elstmtnr])
                x1.add_row(["Existing RNN", cfg.ernntnr])
                x1.add_row(["Existing DBN", cfg.edbntnr])
                x1.add_row(["Existing DNN", cfg.ednntnr])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmtnr, 'Existing LSTM': cfg.elstmtnr,
                        'Existing RNN': cfg.ernntnr, 'Existing DBN': cfg.edbntnr, 'Existing DNN': cfg.ednntnr}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.plot(courses, values, color='plum',marker="o")
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("TNR", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\"+str(cfg.data_name)+" TNR.png")
                plt.show()
                plt.close()

            tnr()

            def trainingtime():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Training Time"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmtrtime),
                    ("Existing LSTM", cfg.lstmtrtime),
                    ("Existing RNN", cfg.rnntrtime),
                    ("Existing DBN", cfg.dbntrtime),
                    ("Existing DNN", cfg.dnntrtime)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Training Time"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "Time in ms"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\Trainingtime.xlsx")
                print("\nTraining Time\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Training Time"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmtrtime])
                x1.add_row(["Existing LSTM", cfg.lstmtrtime])
                x1.add_row(["Existing RNN", cfg.rnntrtime])
                x1.add_row(["Existing DBN", cfg.dbntrtime])
                x1.add_row(["Existing DNN", cfg.dnntrtime])
                print(x1.get_string(title=""))
                plt.show()
                plt.close()

            trainingtime()


            def Detection_Rate():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Detection Rate"),
                    ("Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmdr),
                    ("Existing LSTM", cfg.elstmdr),
                    ("Existing RNN", cfg.ernndr),
                    ("Existing DBN", cfg.edbndr),
                    ("Existing DNN", cfg.ednndr)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Detection Rate"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Classification Algorithms"
                chart.y_axis.title = "TNR (%)"
                ws.add_chart(chart, "E5")
                wb.save("..\Result\\Detection_Rate.xlsx")
                print("\nDetection Rate\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Detection Rate"]
                x1.add_row(["Proposed FedTL-SRSKLSTM ", cfg.pfedtlsrsklstmdr])
                x1.add_row(["Existing LSTM", cfg.elstmdr])
                x1.add_row(["Existing RNN", cfg.ernndr])
                x1.add_row(["Existing DBN", cfg.edbndr])
                x1.add_row(["Existing DNN", cfg.ednndr])
                print(x1.get_string(title=""))

                data = {'Proposed FedTL-SRSKLSTM': cfg.pfedtlsrsklstmdr, 'Existing LSTM': cfg.elstmdr,
                        'Existing RNN': cfg.ernndr, 'Existing DBN': cfg.edbndr, 'Existing DNN': cfg.ednndr}
                courses = list(data.keys())
                values = list(data.values())

                fig = plt.figure(figsize=(10, 5))

                # creating the bar plot
                plt.plot(courses, values, color='cyan',marker="o")
                plt.xlabel('Techniques', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.ylabel('Values (%)', fontweight='bold', fontname="Times New Roman", fontsize=12)
                plt.title("Detection Rate", fontweight='bold', fontname="Times New Roman", fontsize=14)
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.savefig("..\Result\\Detection_Rate.png")
                plt.show()
                plt.close()

            Detection_Rate()


            def CH_fitness_metric():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Proposed ALV-HBA", "Existing HBA", "Existing EHOA", "Existing FOA", "Existing SCSOA"),
                    ("10", 78.231, 73.248, 68.153, 63.106, 57.243),
                    ("20", 83.214, 78.065, 73.113, 68.12, 63.203),
                    ("30", 88.214, 83.124, 78.421, 73.125, 68.142),
                    ("40", 95.452, 88.241, 83.121, 78.032, 73.101),
                    ("50", 101.124, 95.452, 94.245, 88.214, 79.014),
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=6, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Fitness vs. Iteration for CH Selection"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.title = "Iteration"
                chart.y_axis.title = "Fitness"
                ws.add_chart(chart, "E5")
                wb.save("..\\Result\\CHFitnessVs.Iteration.xlsx")
                print("\n(Fitness vs. Iteration for CH Selection)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Proposed ALV-HBA", "Existing HBA", "Existing EHOA", "Existing FOA",
                                  "Existing SCSOA"]
                x1.add_row(["Proposed ALV-HBA", 78.231, 83.214, 88.214, 95.452, 101.124])
                x1.add_row(["Existing HBA", 73.248,78.065, 83.124,88.241, 95.452])
                x1.add_row(["Existing EHOA", 68.153, 73.113, 78.421, 83.121, 94.245])
                x1.add_row(["Existing FOA", 63.106, 68.12, 73.125, 78.032, 88.214])
                x1.add_row(["Existing SCSOA", 57.243, 63.203,68.142, 73.101, 79.014])
                print(x1.get_string(title=""))

                fig = plt.figure()
                tech = ["10", " 20", "30", "40", "50"]
                val = [78.231, 83.214, 88.214, 95.452, 101.124]
                val1 = [ 73.248,78.065, 83.124,88.241, 95.452]
                val2 = [68.153, 73.113, 78.421, 83.121, 94.245]
                val3 = [63.106, 68.12, 73.125, 78.032, 88.214]
                val4 = [57.243, 63.203,68.142, 73.101, 79.014]
                col = ["deeppink", "darkviolet", "navy", "darkslategray", "saddlebrown"]
                plt.plot(tech, val, color="deeppink",marker="*",label="Proposed ALV-HBA")
                plt.plot(tech, val1, color="darkslategray", marker="*",label="Existing HBA")
                plt.plot(tech, val2, color="navy", marker="*",label="Existing EHOA")
                plt.plot(tech, val3, color="saddlebrown", marker="*",label="Existing FOA")
                plt.plot(tech, val4, color="darkviolet", marker="*",label="Existing SCSOA")

                plt.legend()
                plt.xticks(fontname="Times New Roman")
                plt.yticks(fontname="Times New Roman")
                plt.xlabel('Iterations', fontname='Times New Roman', weight="bold")
                plt.ylabel('Fitness', fontname='Times New Roman', weight="bold")
                plt.title('Fitness vs. Iteration for CH Selection', fontname='Times New Roman', weight="bold")
                plt.savefig("..\\Result\\CHFitnessVsiteration.png")
                plt.show()
                plt.close()

            CH_fitness_metric()
            self.btn_result_graphs.config(state="disable")
            messagebox.showinfo("Info Message", "Generate Tables And Graphs was done successfully...")
        else:
            messagebox.showinfo("Info Message","Please done the Data Classification Testing or Attack Classification Testing First ...")

    def clear(self):
        self.data_textarea_process.configure(state="normal")
        self.data_textarea_result.configure(state="normal")
        self.data_textarea_process.delete("1.0", "end")
        self.data_textarea_result.delete("1.0", "end")
        self.btn_tr_dataset.config(state="normal")
        self.btn_DDR.config(state="normal")
        self.btn_Numeralization.config(state="normal")
        self.btn_Normalization.config(state="normal")
        self.btn_data_balancing.config(state="normal")
        self.btn_feature_extraction.config(state="normal")
        self.btn_feature_evaluation.config(state="normal")
        self.btn_dimensionality_reduction.config(state="normal")
        self.btn_dataset_splitting.config(state="normal")
        self.btn_training.config(state="normal")
        self.btn_testing.config(state="normal")
        self.btn_initalize_nodes.config(state="normal")
        self.btn_ch_selection.config(state="normal")
        self.btn_clustering.config(state="normal")
        self.btn_testing.config(state="normal")
        self.btn_initalize_nodes.config(state="normal")
        self.btn_ch_selection.config(state="normal")
        self.btn_clustering.config(state="normal")
        self.btn_map_reduce.config(state="normal")
        self.btn_sink_node.config(state="normal")
        self.btn_feature_similarity.config(state="normal")
        self.btn_ts_feature_extraction.config(state="normal")
        self.btn_classifier.config(state="normal")
        self.dataset_cb.config(state="normal")
        self.cmb_nodes.config(state="normal")
        self.dataset_cb.set('')
        self.cmb_nodes.set('')


    def exit(self):
        self.root.destroy()

def getListOfFiles(image_folder):
    listOfFile = os.listdir(image_folder)
    allFiles = list()
    for entry in listOfFile:
        fullPath = os.path.join(image_folder, entry)
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)
    return allFiles

root = Tk()
root.title("INTRUSION DETECTION SYSTEM")
root.geometry("1150x700")
root.resizable(0, 0)
root.configure(bg="wheat")
od = Main_GUI(root)
root.mainloop()